"""Inspect both spreadsheets: list sheets, sizes, and dump non-empty content."""
import json
from pathlib import Path
from openpyxl import load_workbook

SRC_FECHAMENTO = Path("/home/nandoravioli/bia4u/rumo/Copy of Fechamento MBC 02.2026.xlsx")
SRC_AUDIT = Path("/home/nandoravioli/bia4u/rumo/MBC_formula_audit_v2.xlsx")
OUT_DIR = Path("/home/nandoravioli/bia4u/rumo/work/analysis")
OUT_DIR.mkdir(parents=True, exist_ok=True)


def summarize(path: Path, label: str):
    print(f"\n=== {label} :: {path.name} ===")
    wb = load_workbook(str(path), data_only=False, read_only=False)
    summary = {}
    for ws in wb.worksheets:
        summary[ws.title] = {
            "max_row": ws.max_row,
            "max_col": ws.max_column,
        }
        print(f"  {ws.title}: rows={ws.max_row}, cols={ws.max_column}")
    (OUT_DIR / f"{label}_sheets.json").write_text(json.dumps(summary, indent=2, ensure_ascii=False))
    return wb


wb_f = summarize(SRC_FECHAMENTO, "fechamento")
wb_a = summarize(SRC_AUDIT, "audit")
