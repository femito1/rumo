"""Extract all the values the workbook expects to receive externally.

The model is: Base_Resultado Mensal_V2 is the operational input sheet. Every row
that has a label in column A and a number in column C (Jan) and/or D (Feb) is a
manually-entered value that today is typed in by hand. Those are the targets we
need to pull from Legal Manager.

We also flag the hardcoded numbers found inside formulas in the area sheets,
since those are "phantom" inputs (numeric literals embedded in formulas that
should be replaced by API data).
"""
import csv
import json
import re
from pathlib import Path
from openpyxl import load_workbook

SRC = Path("/home/nandoravioli/bia4u/rumo/Copy of Fechamento MBC 02.2026.xlsx")
OUT = Path("/home/nandoravioli/bia4u/rumo/work/analysis/manual_inputs.tsv")
OUT_JSON = Path("/home/nandoravioli/bia4u/rumo/work/analysis/manual_inputs.json")
HARDCODES = Path("/home/nandoravioli/bia4u/rumo/work/analysis/hardcoded_in_formulas.tsv")

wb_v = load_workbook(str(SRC), data_only=True)
wb_f = load_workbook(str(SRC), data_only=False)

base_v = wb_v["Base_Resultado Mensal_V2"]
base_f = wb_f["Base_Resultado Mensal_V2"]

manual_rows = []
for r in range(3, base_v.max_row + 1):
    label = base_v.cell(r, 1).value
    if label is None:
        continue
    label = str(label).strip()
    if not label:
        continue
    jan_v = base_v.cell(r, 3).value
    feb_v = base_v.cell(r, 4).value
    jan_f = base_f.cell(r, 3).value
    feb_f = base_f.cell(r, 4).value

    def is_manual(formula, value):
        if value in (None, "", 0):
            return False
        if formula is None:
            return False
        if isinstance(formula, (int, float)):
            return True
        s = str(formula)
        if s.startswith("="):
            return False
        return True

    jan_manual = is_manual(jan_f, jan_v)
    feb_manual = is_manual(feb_f, feb_v)
    if not (jan_manual or feb_manual):
        continue

    manual_rows.append({
        "row": r,
        "label": label,
        "jan_2025_value": jan_v,
        "jan_2025_formula": jan_f,
        "jan_is_manual": jan_manual,
        "feb_2025_value": feb_v,
        "feb_2025_formula": feb_f,
        "feb_is_manual": feb_manual,
    })

with open(OUT, "w", newline="", encoding="utf-8") as f:
    w = csv.writer(f, delimiter="\t")
    w.writerow([
        "row", "label", "jan_value", "jan_manual", "feb_value", "feb_manual",
    ])
    for r in manual_rows:
        w.writerow([r["row"], r["label"], r["jan_2025_value"], r["jan_is_manual"],
                    r["feb_2025_value"], r["feb_is_manual"]])

OUT_JSON.write_text(json.dumps(manual_rows, indent=2, ensure_ascii=False, default=str))

# Also collect hardcoded numeric literals inside formulas of area / DRE sheets
literal_re = re.compile(r"(?<![A-Za-z_!:$])\d{2,7}(?:\.\d+)?(?![A-Za-z_])")
target_sheets = ["Areas Sintetico atualizado", "DRE 2026", "Rateio Mensal", "Amortização", "Resumo_Recebidas 2025_2026"]
hardcodes = []
for sn in target_sheets:
    if sn not in wb_f.sheetnames:
        continue
    ws_f = wb_f[sn]
    ws_v = wb_v[sn]
    for row in ws_f.iter_rows():
        for cell in row:
            v = cell.value
            if not isinstance(v, str) or not v.startswith("="):
                continue
            literals = literal_re.findall(v[1:])
            literals = [x for x in literals if float(x) >= 100]
            if not literals:
                continue
            cached = ws_v.cell(cell.row, cell.column).value
            hardcodes.append({
                "sheet": sn,
                "cell": cell.coordinate,
                "formula": v,
                "literals": literals,
                "cached_value": cached,
            })

with open(HARDCODES, "w", newline="", encoding="utf-8") as f:
    w = csv.writer(f, delimiter="\t")
    w.writerow(["sheet", "cell", "formula", "literals", "cached_value"])
    for h in hardcodes:
        w.writerow([h["sheet"], h["cell"], h["formula"], "|".join(h["literals"]), h["cached_value"]])

print(f"Manual rows: {len(manual_rows)} -> {OUT}")
print(f"Formulas with embedded literals: {len(hardcodes)} -> {HARDCODES}")
