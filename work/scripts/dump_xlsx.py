"""Dump worksheets to readable TSV/JSON for analysis."""
import csv
import json
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

SRC_FECHAMENTO = Path("/home/nandoravioli/bia4u/rumo/Copy of Fechamento MBC 02.2026.xlsx")
SRC_AUDIT = Path("/home/nandoravioli/bia4u/rumo/MBC_formula_audit_v2.xlsx")
OUT_DIR = Path("/home/nandoravioli/bia4u/rumo/work/analysis")


def dump_sheets_tsv(path: Path, label: str, data_only: bool):
    suffix = "_values" if data_only else "_formulas"
    out_dir = OUT_DIR / f"{label}{suffix}"
    out_dir.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(str(path), data_only=data_only)
    for ws in wb.worksheets:
        safe = "".join(c if c.isalnum() or c in "._-" else "_" for c in ws.title).strip("_")
        fp = out_dir / f"{safe}.tsv"
        with open(fp, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f, delimiter="\t", quoting=csv.QUOTE_MINIMAL)
            for r_idx, row in enumerate(ws.iter_rows(values_only=False), start=1):
                cells = []
                for cell in row:
                    v = cell.value
                    if v is None:
                        cells.append("")
                    else:
                        s = str(v).replace("\t", " ").replace("\n", " | ")
                        cells.append(s)
                w.writerow(cells)


for d in (True, False):
    dump_sheets_tsv(SRC_FECHAMENTO, "fechamento", d)
dump_sheets_tsv(SRC_AUDIT, "audit", True)
print("Dumped TSVs.")
