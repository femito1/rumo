"""Per-person / per-account reconciliation of per-área Custo equipe, Jan–Abr 2026.

Why this exists, and why it is NOT build_janabr_diff.py
-------------------------------------------------------
`build_janabr_diff.py` proves its *despesa* attribution structurally: the ten
institutional families ARE the components of `r198`, so their deltas must sum to the
Despesas Indiretas delta. That is an accounting identity and it leaves no residue.

Its *custo equipe* attribution was weaker, and this script is what closed that hole.
§1–§4 of that document are four tables, each showing one cause in isolation; nothing ever
added them up and compared against the per-área delta. Doing that by hand left a residual
(Econômico Jan/Fev ~ -3.000 was covered by none of the four), and §3's list of
"lançamentos avulsos" was HAND-PICKED — it named rows 34/35/36/43/47/51/54/55 while the
Contencioso and Arbitragem blocks have their own one-off rows it never mentioned.

So: no hand-picked row lists here. Both sides are decomposed to (person, account) and
every difference is bucketed into a NAMED cause. The residual is printed per área/month
and is the honest measure of what is still unexplained. It is currently 0,00 everywhere.

What it found (2026-08-03), neither of which was in the docs:

* **ISS trimestral (030.010.0160) is a PRESENTATION difference, not an error.** In
  jan/abr it posts per lawyer (382,16 or 507,14 each) and we fold it into that lawyer's
  área; the book types it as ONE área-level "ISS Trimestral" row (r25/r54/r79). Same
  total, so it cancels at área level — but it makes every single lawyer's line differ,
  which is why a per-person read looks alarming until you net it.
* **The convênio difference in jan/fev is the SISJURI memo itself, not our parse.** The
  jan/fev memos state a different health-plan base from mar–jun and are internally
  consistent at that base: EHF "1.795,86-1.192,36 (Parte MBC)=603,50" versus March's
  "3.520,31 - 1.956,21 (Parte MBC) = 1.564,10". Both parse correctly; we faithfully
  report what finance wrote. The BOOK, by contrast, types the SAME 1.564,10 / 2.526,09 in
  all six months, so in jan/fev it does not follow its own memo. Worth 2.962,41 per month
  on Econômico and +1.911,95 on Arbitragem in February.
  ⚠ I first called this a regex bug ("we take the last number"). That was WRONG — the
  regex anchors on "(Parte MBC) =" and both months match it correctly. Check whether the
  SOURCE data differs before blaming the parser.

Run: cd backend && python -m scripts.reconcile_custo_equipe
"""
from __future__ import annotations

import re
import unicodedata
from collections import defaultdict
from pathlib import Path
from typing import Any

import openpyxl

REPO = Path(__file__).resolve().parents[2]
WORKBOOK = REPO / "reference" / "workbook" / "Fechamento MBC 06.2026.xlsx"

MESES = {1: "Janeiro", 2: "Fevereiro", 3: "Março", 4: "Abril", 5: "Maio", 6: "Junho"}
BASE_COL = {1: 3, 2: 4, 3: 5, 4: 6, 5: 7, 6: 8}

#: (header row, last leaf row, our área name). The ranges are the header cells' OWN SUM
#: formulas, read from the workbook rather than guessed: r5 =SUM(6:27), r30 =SUM(31:57),
#: r60 =SUM(61:79). Guessing wider ranges silently pulls in rows the book excludes from
#: its own total (r28/r29 Participação+Repasse, r58/r59, r80/r81) and a residual then
#: appears to come from nowhere.
BLOCKS: tuple[tuple[int, int, str], ...] = (
    (5, 27, "Contencioso"),
    (30, 57, "Econômico"),
    (60, 79, "Arbitragem"),
)

ISS = "030.010.0160"
CONVENIO = "030.010.0110"
VALE = "030.010.0100/0220"

#: Book leaf label -> the account it corresponds to on our side. Matched on a normalised
#: substring of the part AFTER the person's name.
LABEL_TO_ACCOUNT: tuple[tuple[str, str], ...] = (
    ("convenio medico", CONVENIO),
    ("convenio med", CONVENIO),
    ("distribuicao mensal", "030.010.0010"),
    ("reajuste de distribuicao", "030.010.0010"),
    ("pro labore", "030.010.0130"),
    ("pro labores", "030.010.0130"),
    ("bolsa auxilio", "030.010.0140"),
    ("aasp", "030.010.0150"),
    ("iss trimestral", ISS),
    ("iss - trimestral", ISS),
    ("vale refeicao", VALE),
    ("vale transporte", VALE),
)

#: Área-level (non-person) leaf labels inside a Custo equipe block. These are typed as one
#: row for the whole área where our side is per person, so they are netted at área level.
#: NOTE "seguro de vida" is deliberately NOT here: the book has BOTH a per-person variant
#: (r53 "Vitoria Santos de Rezende Seguro de vida") and a bare one (r55), and the bare row
#: still sits inside the block's own SUM — treating it as área-level dropped it from the
#: comparison entirely and left an unexplained -92,44 in March.
_AREA_LEVEL = (
    "iss trimestral",
    "iss - trimestral",
    "vale refeicao",
    "vale transporte",
    "participacao comissao",
    "participacao/comissao",
    "repasse",
)


def _norm(s: object) -> str:
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"[^a-z0-9]+", " ", s.lower()).strip()


def _brl(v: float | None) -> str:
    if v is None:
        return "—"
    s = f"{abs(v):,.2f}".replace(",", "@").replace(".", ",").replace("@", ".")
    return f"{'-' if v < 0 else ''}{s}"


def _split_leaf(label: str) -> tuple[str, str | None]:
    """Split a book leaf label into (person, account).

    The account is identified by a known label fragment; the person is whatever precedes
    it. Deriving the person this way (rather than splitting on " - ") is what makes the
    no-dash leaves work — e.g. "Vitoria Santos de Rezende Bolsa auxilio / IR" — and it
    keeps the FULL name, which the sigla matcher needs (truncating "isabel de almeida rego
    campinho" to three tokens drops the C that IAC matches on).
    """
    nl = _norm(label)
    best: tuple[int, str, str] | None = None
    for frag, acct in LABEL_TO_ACCOUNT:
        i = nl.find(frag)
        if i >= 0 and (best is None or i < best[0]):
            best = (i, frag, acct)
    if best is not None:
        return nl[: best[0]].strip(), best[2]
    # No known account fragment. If the label still has the "Nome - resto" shape it is a
    # person's row under a heading we do not model ("Subsidio de Pos-graduação", or r71
    # where the account name was simply never typed). Return the person with an UNKNOWN
    # account so it nets at person level. Without this the whole label becomes a phantom
    # person and its value shows up twice — once as "só no livro" and once as a bogus
    # per-account delta.
    raw = str(label)
    if " - " in raw:
        return _norm(raw.split(" - ")[0]), "?"
    return nl, None


def match_sigla(sigla: str, person: str) -> bool:
    """Is ``sigla`` an ordered subsequence of the person's word initials?

    The snapshot carries no full names, so the sigla IS the join key. SISJURI siglas are
    not "first N initials" (AM = Aurelio Marchini Santos, IAC = Isabel de Almeida rego
    Campinho, RB = Ricardo franco Botelho), but every one observed is an ordered
    subsequence of the initials with the first letter matching. Callers require the hit to
    be unique within the block, so an ambiguous pair is reported, never silently absorbed.
    """
    initials = [w[0] for w in person.split() if w]
    want = list(sigla.lower())
    if not initials or not want or initials[0] != want[0]:
        return False
    i = 0
    for ch in initials:
        if i < len(want) and ch == want[i]:
            i += 1
    return i == len(want)


def our_rows(snap: dict[str, Any]) -> list[dict[str, Any]]:
    """The exact row set ``dre.py`` feeds to ``derive_area_custo_equipe``."""
    from app.closing.dre import is_adm_grupo

    home = {str(k): str(v) for k, v in (snap.get("home_area") or {}).items()}
    vale = [
        v
        for v in (snap.get("custo_equipe_area") or [])
        if not is_adm_grupo(home.get(str(v.get("sigla") or "")))
    ]
    return [*(snap.get("custo_equipe_deriv") or []), *vale]


def our_overrides(snap: dict[str, Any]) -> dict[str, Any]:
    """The convênio overrides, applying the SAME stale-memo guard as ``dre.py``.

    This used to be a private copy of the override loop, and when the guard landed in
    ``dre.py`` this script silently kept the old behaviour — so the reconciliation
    reported numbers the product no longer produced. Reusing
    ``_memo_describes_this_month`` is what keeps the two from drifting again.
    """
    from app.closing.custo_equipe_deriv import CONVENIO_ACCOUNT, LawyerOverride
    from app.closing.dre import _memo_describes_this_month

    posted: dict[str, float] = {}
    for row in snap.get("custo_equipe_deriv") or []:
        if str(row.get("id_conta") or "") == CONVENIO_ACCOUNT:
            sg = str(row.get("sigla") or "").strip()
            if sg:
                posted[sg] = round(posted.get(sg, 0.0) + float(row.get("valor") or 0.0), 2)

    out: dict[str, Any] = {}
    for memo in snap.get("convenio_memo") or []:
        sigla = str(memo.get("sigla") or "").strip()
        parsed = memo.get("parsed_valor")
        if not sigla or parsed is None:
            continue
        if not _memo_describes_this_month(str(memo.get("raw_memo") or ""), posted.get(sigla)):
            continue
        out[sigla] = LawyerOverride(set_account={CONVENIO_ACCOUNT: float(parsed)})
    return out


def our_by_person_account(snap: dict[str, Any]) -> tuple[
    dict[str, dict[tuple[str, str], float]], dict[str, float]
]:
    """área -> (sigla, account) -> value, plus the production per-área totals.

    Per-person values come from the PRODUCTION functions, called one lawyer at a time
    (the derivation is a sum over independent lawyers), so this breakdown cannot drift
    from the total it explains.
    """
    from app.closing.custo_equipe_deriv import build_area_splits, derive_area_custo_equipe

    home = {str(k): str(v) for k, v in (snap.get("home_area") or {}).items()}
    splits = build_area_splits(snap.get("rateio_grupo") or [], home)
    overrides = our_overrides(snap)
    rows = our_rows(snap)

    per: dict[str, dict[tuple[str, str], float]] = defaultdict(lambda: defaultdict(float))
    siglas = {str(r.get("sigla") or "").strip() for r in rows} - {""}
    for sg in sorted(siglas):
        mine = [r for r in rows if str(r.get("sigla") or "").strip() == sg]
        accts = {str(r.get("id_conta") or "") for r in mine}
        ov = {sg: overrides[sg]} if sg in overrides else None
        # One account at a time so the breakdown is per (person, account). The convênio
        # override must ride with its own account only.
        for acct in sorted(accts):
            sub = [r for r in mine if str(r.get("id_conta") or "") == acct]
            sub_ov = ov if (ov and acct == CONVENIO) else None
            for area, val in derive_area_custo_equipe(sub, splits, overrides=sub_ov).items():
                if abs(val) > 0.005:
                    per[area][(sg, acct)] += round(val, 2)
        if ov and CONVENIO not in accts:
            for area, val in derive_area_custo_equipe([], splits, overrides=ov).items():
                if abs(val) > 0.005:
                    per[area][(sg, CONVENIO)] += round(val, 2)

    totals = derive_area_custo_equipe(rows, splits, overrides=overrides)
    return (
        {a: {k: round(v, 2) for k, v in d.items()} for a, d in per.items()},
        {a: round(v, 2) for a, v in totals.items()},
    )


def book_by_person_account(
    base: Any, month: int
) -> dict[str, tuple[dict[tuple[str, str], float], list[tuple[int, str, float]]]]:
    """área -> ((person, account) -> value, [unmapped leaf rows])."""
    out: dict[str, tuple[dict[tuple[str, str], float], list[tuple[int, str, float]]]] = {}
    for head, end, area in BLOCKS:
        per: dict[tuple[str, str], float] = defaultdict(float)
        odd: list[tuple[int, str, float]] = []
        for r in range(head + 1, end + 1):
            lab = base.cell(r, 1).value
            if lab is None:
                continue
            raw = base.cell(r, BASE_COL[month]).value
            val = float(raw) if isinstance(raw, (int, float)) else 0.0
            if abs(val) < 0.005:
                continue
            nl = _norm(lab)
            person, acct = _split_leaf(str(lab))
            if any(nl.startswith(p) for p in _AREA_LEVEL):
                per[("(área)", acct or "?")] += val
                continue
            if acct is None and person:
                # The book has a leaf whose account name was never typed — r71 is
                # literally "Joao Gabriel Previero de Arruda Sampaio - " (Pró-Labore, by
                # value and by position after r70 Distribuição). Attribute it to the
                # person with an UNKNOWN account so it still nets against our side
                # instead of masquerading as an unexplained residual.
                per[(person, "?")] += val
                continue
            if acct is None or not person:
                odd.append((r, str(lab).strip(), val))
                continue
            per[(person, acct)] += val
        out[area] = ({k: round(v, 2) for k, v in per.items()}, odd)
    return out


def main() -> None:
    from dotenv import load_dotenv

    load_dotenv(REPO / "backend" / ".env")
    from app.api.providers import get_snapshot_store

    snaps = get_snapshot_store().snapshots_by_year(2026, client_id="mbc")
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
    base = wb["Base_Resultado Mensal_V2"]

    grand: dict[str, float] = defaultdict(float)
    for m in MESES:
        snap = snaps[m]
        ours, totals = our_by_person_account(snap)
        book = book_by_person_account(base, m)
        print(f"\n{'#' * 78}\n# {MESES[m]} 2026\n{'#' * 78}")
        for head, _end, area in BLOCKS:
            book_tot = round(float(base.cell(head, BASE_COL[m]).value or 0.0), 2)
            our_tot = round(totals.get(area, 0.0), 2)
            delta = round(our_tot - book_tot, 2)
            bk, odd = book[area]
            print(
                f"\n{area}: nosso {_brl(our_tot)} · planilha {_brl(book_tot)} · Δ {_brl(delta)}"
            )

            # Map our siglas to book persons (unique hit required).
            people = {p for p, _a in bk} - {"(área)"}
            sig_to_person: dict[str, str] = {}
            for sg in sorted({s for s, _a in ours.get(area, {})}):
                hits = [p for p in people if match_sigla(sg, p)]
                if len(hits) == 1:
                    sig_to_person[sg] = hits[0]

            buckets: dict[str, float] = defaultdict(float)
            unexplained: list[str] = []

            # 1. ISS: ours is per-person, the book types one área-level row.
            our_iss = sum(v for (_s, a), v in ours.get(area, {}).items() if a == ISS)
            bk_iss = sum(v for (_p, a), v in bk.items() if a == ISS)
            if abs(our_iss) > 0.005 or abs(bk_iss) > 0.005:
                buckets["ISS trimestral (nosso por pessoa × livro por área)"] += round(
                    our_iss - bk_iss, 2
                )

            # 2. Vale: ours per-person, book área-level.
            our_vale = sum(v for (_s, a), v in ours.get(area, {}).items() if a == VALE)
            bk_vale = sum(v for (_p, a), v in bk.items() if a == VALE)
            if abs(our_vale) > 0.005 or abs(bk_vale) > 0.005:
                buckets["Vale advogados (regra do cliente: sempre incluir)"] += round(
                    our_vale - bk_vale, 2
                )

            # 3. Everything else, per (person, account).
            for sg in sorted({s for s, _a in ours.get(area, {})}):
                if sg not in sig_to_person:
                    for (s, a), v in ours.get(area, {}).items():
                        if s == sg and a not in (ISS, VALE) and abs(v) > 0.005:
                            unexplained.append(f"{sg} {a} nosso {_brl(v)} — SEM MATCH no livro")
            for sg, person in sorted(sig_to_person.items()):
                accts = {a for s, a in ours.get(area, {}) if s == sg and a not in (ISS, VALE)} | {
                    a for p, a in bk if p == person and a not in (ISS, VALE)
                }
                # A leaf the book left unlabelled ("?") can only be netted against the
                # person's total, so compare person-level when one is present.
                if "?" in accts:
                    o = round(
                        sum(
                            v
                            for (s, a), v in ours.get(area, {}).items()
                            if s == sg and a not in (ISS, VALE)
                        ),
                        2,
                    )
                    b = round(
                        sum(v for (p, a), v in bk.items() if p == person and a not in (ISS, VALE)),
                        2,
                    )
                    d = round(o - b, 2)
                    if abs(d) > 0.005:
                        buckets[f"{sg}: total da pessoa (livro tem linha sem conta)"] += d
                    continue
                for acct in sorted(accts):
                    o = ours.get(area, {}).get((sg, acct), 0.0)
                    b = bk.get((person, acct), 0.0)
                    d = round(o - b, 2)
                    if abs(d) < 0.005:
                        continue
                    if acct == CONVENIO:
                        buckets["Convênio: memo do mês (jan/fev) × constante do livro"] += d
                    elif acct == "030.010.0150":
                        buckets["AASP: livro em Custo equipe, DB em Despesas Área"] += d
                    else:
                        buckets[f"{sg} {acct}"] += d
            # Book persons with no counterpart on our side.
            for person in sorted(people - set(sig_to_person.values())):
                tot = round(
                    sum(v for (p, a), v in bk.items() if p == person and a not in (ISS, VALE)), 2
                )
                if abs(tot) > 0.005:
                    buckets[f"só no livro: {person}"] += -tot
            # Book leaves whose label maps to no account of ours.
            for row, lab, val in odd:
                buckets[f"só no livro r{row}: {lab[:34]}"] += -val

            for name, val in sorted(buckets.items(), key=lambda kv: -abs(kv[1])):
                if abs(val) > 0.005:
                    print(f"    {_brl(val):>12}  {name}")
                    grand[name] += val
            for u in unexplained:
                print(f"    {'?':>12}  {u}")
            resid = round(delta - sum(buckets.values()), 2)
            flag = "OK" if abs(resid) < 0.02 else "<<<< NÃO EXPLICADO"
            print(f"    {'-' * 60}\n    {_brl(resid):>12}  RESÍDUO  {flag}")

    print(f"\n{'#' * 78}\n# Jan–Abr somado, por causa\n{'#' * 78}")
    for name, val in sorted(grand.items(), key=lambda kv: -abs(kv[1])):
        if abs(val) > 0.005:
            print(f"  {_brl(val):>12}  {name}")


if __name__ == "__main__":
    main()
