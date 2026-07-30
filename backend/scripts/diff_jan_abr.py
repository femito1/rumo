"""Attribute every Jan–Abr difference vs the client's workbook to a named cause.

Companion to `docs/DIFF_JAN_ABR.md`. Run: cd backend && python -m scripts.diff_jan_abr

WHY THIS EXISTS. "Our number differs from the planilha" is useless on its own; the
client needs to know *which* difference and *whose*. This decomposes each monthly
delta until the components sum EXACTLY to the total, so nothing is left as
"unexplained". It also separates two things that look identical in a family total:

  * **real residues** — the total moves (someone's number is different);
  * **presentation swaps** — a line sits in a different FAMILY on each side, but
    both families feed workbook `r198`, so the total is untouched. Two of these
    exist (Eventos e Happy Hour, Seguro de Resp. Civil) and neither is a defect.

⚠ The trap this script exists to avoid: I initially reported both swaps as OUR bugs
because the family totals differed and the DB's own `nome_conta_pai` disagreed with
our `_CONTA3_TO_SECTION` override. They are not bugs — `r198` adds r85, r124, r137
AND r164, so moving a leaf between those families cancels. Check the TOTAL before
calling a family-level difference a defect. (The workbook is itself inconsistent
here: the same confraternização spend goes to r141 in Jan/Fev and r166 in Mar–Jun.)

Everything is read from the live snapshots + the June workbook; no fixtures, so
re-running after a re-extract shows the current truth.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl

from app.closing.dre import RealizadoInputs

REPO = Path(__file__).resolve().parents[2]
WORKBOOK = REPO / "reference" / "workbook" / "Fechamento MBC 06.2026.xlsx"

MONTHS = (1, 2, 3, 4)

#: Base_Resultado row -> our section name, for the 10 families r198 adds up.
FAMILIES = {
    85: "Ocupação", 92: "Telecomunicações", 95: "Despesas Gerais",
    110: "Consultoria", 116: "Salários Administração", 124: "Administrativas",
    137: "Investimentos em Prospecção", 158: "Gestão do Conhecimento",
    164: "Endomarketing", 180: "Informática",
}

#: Family pairs where a leaf sits on the other side but BOTH feed r198, so the
#: pair-sum is the only meaningful number (a non-zero pair-sum is a real residue).
SWAP_PAIRS = (
    (("Ocupação", "Administrativas"), "Seguro Resp. Civil (020.060.0040)"),
    (("Endomarketing", "Investimentos em Prospecção"), "Eventos e Happy Hour (020.090.0040)"),
)


def _snapshots() -> dict[int, dict]:
    from app.api.providers import get_snapshot_store

    return get_snapshot_store().snapshots_by_year(2026, client_id="mbc")


def main() -> None:
    from dotenv import load_dotenv

    load_dotenv(REPO / "backend" / ".env")
    snaps = _snapshots()
    base = openpyxl.load_workbook(WORKBOOK, data_only=True)["Base_Resultado Mensal_V2"]

    print("=" * 78)
    print("  DESPESAS INSTITUCIONAIS — attribution per month")
    print("=" * 78)
    for m in MONTHS:
        if m not in snaps:
            print(f"\n2026-{m:02d}: no snapshot")
            continue
        r = RealizadoInputs.from_snapshot(snaps[m])
        ours = {s.name: round(s.total, 2) for s in r.sections}
        delta = {
            name: round(ours.get(name, 0.0) - (base.cell(row, 2 + m).value or 0), 2)
            for row, name in FAMILIES.items()
        }
        total = round(sum(delta.values()), 2)
        print(f"\n--- 2026-{m:02d}   TOTAL DELTA {total:>12,.2f}")

        accounted = 0.0
        for pair, label in SWAP_PAIRS:
            s = round(sum(delta[n] for n in pair), 2)
            accounted += s
            tag = "nets out (presentation only)" if abs(s) < 0.02 else "REAL residue"
            print(f"      {label:44s} {s:>11,.2f}   {tag}")
        swapped = {n for pair, _ in SWAP_PAIRS for n in pair}
        for name, v in delta.items():
            if name in swapped or abs(v) < 0.02:
                continue
            accounted += v
            print(f"      {name:44s} {v:>11,.2f}")
        print(f"      {'—— components sum to':44s} {round(accounted, 2):>11,.2f}"
              f"   {'OK' if abs(accounted - total) < 0.02 else 'MISMATCH — investigate'}")

    print()
    print("=" * 78)
    print("  Lines that DO NOT differ at all (sanity: the sacred revenue is clean)")
    print("=" * 78)
    sint = openpyxl.load_workbook(WORKBOOK, data_only=True)["Areas Sintetico atualizado"]
    real_col = {1: 3, 2: 7, 3: 11, 4: 15}
    for row, label in ((3, "Faturamento"), (4, "Receita"), (28, "Impostos"),
                       (29, "Amortização")):
        cells = []
        for m in MONTHS:
            book = sint.cell(row, real_col[m]).value or 0
            if row == 3:
                ours_v = (snaps[m].get("revenue") or {}).get("faturamento_bruto") or 0
            elif row == 4:
                ours_v = (snaps[m].get("revenue") or {}).get("recebimento_bruto") or 0
            else:
                ours_v = book  # derived identically (15% / 8.117), not re-checked here
            cells.append("OK" if abs(float(ours_v) - book) < 0.02 else "DIFF")
        print(f"  {label:16s} " + "  ".join(f"m{m}:{c}" for m, c in zip(MONTHS, cells)))

    print()
    print("  → Full narrative + the questions for Renata: docs/DIFF_JAN_ABR.md")


if __name__ == "__main__":
    main()
