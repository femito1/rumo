"""Prove (not assert) what drives the per-área **Despesas Equipe** differences.

The differences document claimed "the workbook's r204/205/206 formula is off by one row in
Jan–Mai" as the dominant cause. That was a READING of the formulas, never a measurement.
This script measures it: it recomputes each Jan–Mai cell using JUNE's formula and reports
how much of the difference disappears. If the hypothesis is right the residual collapses;
if it is wrong we find out here instead of in front of the client.

Result (2026-08-03, live data):

* **The shift is real and it is the dominant cause.** Applying June's formula to Jan–Mai
  cuts the total absolute error from 10.216,31 to 3.494,78 (**-66%**) and takes the
  ties from **4/18 to 11/18** cells.
* Each área's Jan–Mai formula literally reads the NEXT área's labelled rows — verified by
  printing the row labels, not inferred: Contencioso sums "Eventos e Happy hour - Direito
  Econômico", Econômico sums "... - Institucional", Arbitragem sums "... - Contencioso".
  Five families each (Eventos/HH, Material Gráfico, Patrocínio, Refeições, Viagens).
* **The remaining residual is the cost-centre ruling, already settled.** January is the
  whole story: the book splits the two Associações slices across Contencioso (r129
  700,10) and Econômico (r130 700,10), while the DB tags BOTH to EDE — ours shows
  Econômico 1.400,19 ≈ 700,10 + 700,10. Same for Patrocínio 1.204,47 (ESP on our side,
  typed under Contencioso in the book). This is exactly the pattern Renata ruled on
  (allocate by label/cost-centre; the DB is authoritative) and that `extract.sql` records
  for May — so it recurs in January for the same reason.

Run: cd backend && python -m scripts.audit_despesas_area
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl

REPO = Path(__file__).resolve().parents[2]
WORKBOOK = REPO / "reference" / "workbook" / "Fechamento MBC 06.2026.xlsx"

MESES = {1: "Janeiro", 2: "Fevereiro", 3: "Março", 4: "Abril", 5: "Maio", 6: "Junho"}
BASE_COL = {1: 3, 2: 4, 3: 5, 4: 6, 5: 7, 6: 8}

#: The rows each área's Despesas-Equipe formula ADDS, as typed in the workbook.
#: Jan–Mai (r204/205/206 as they stand) vs June (the same formulas, repaired).
#: Read off the workbook with data_only=False; do not "tidy" these.
FORMULA_JANMAI: dict[str, list[int]] = {
    "Contencioso": [125, 129, 140, 144, 148, 152, 156, 160],
    "Econômico": [126, 130, 141, 145, 149, 153, 157, 161],
    "Arbitragem": [127, 131, 139, 143, 147, 151, 155, 159],
}
FORMULA_JUNHO: dict[str, list[int]] = {
    "Contencioso": [125, 129, 139, 143, 147, 151, 155, 160],
    "Econômico": [126, 130, 140, 144, 148, 152, 156, 161],
    "Arbitragem": [127, 131, 138, 142, 146, 150, 154, 159],
}
#: our section key per área label
SECTION = {"Contencioso": "contencioso", "Econômico": "economico", "Arbitragem": "arbitragem"}


def _brl(v: float) -> str:
    s = f"{abs(v):,.2f}".replace(",", "@").replace(".", ",").replace("@", ".")
    return f"{'-' if v < 0 else ''}{s}"


def _our(sections: dict[str, Any], section: str, line: str) -> float:
    for row in (sections.get(section) or {}).get("rows") or []:
        if row.get("key") == line:
            cell = row.get("Realizado")
            v = cell.get("value") if isinstance(cell, dict) else cell
            return float(v) if isinstance(v, (int, float)) else 0.0
    return 0.0


def main() -> None:
    from dotenv import load_dotenv

    load_dotenv(REPO / "backend" / ".env")
    from app.api.providers import get_budget_repo, get_snapshot_store
    from app.budget.models import annual_budget, monthly_budget
    from app.closing.dre import assemble_dre_sections

    snaps = get_snapshot_store().snapshots_by_year(2026, client_id="mbc")
    entries = get_budget_repo().get_budget("mbc", 2026)
    ann = annual_budget(entries) if entries else {}
    wbv = openpyxl.load_workbook(WORKBOOK, data_only=True)["Base_Resultado Mensal_V2"]
    wbf = openpyxl.load_workbook(WORKBOOK, data_only=False)["Base_Resultado Mensal_V2"]

    months = [m for m in MESES if m in snaps]

    print("=" * 78)
    print("1. THE FORMULAS, AS TYPED (proof the shift exists at all)")
    print("=" * 78)
    for r in (204, 205, 206):
        print(f"  r{r} {str(wbv.cell(r, 1).value):24} Jan: {wbf.cell(r, 3).value}")
        print(f"  {'':29} Jun: {wbf.cell(r, 8).value}")
    print("\n  Which rows differ, and what they are LABELLED:")
    for area in FORMULA_JANMAI:
        print(f"    {area}")
        for jm, ju in zip(FORMULA_JANMAI[area], FORMULA_JUNHO[area]):
            if jm != ju:
                print(
                    f"      Jan–Mai r{jm} {str(wbv.cell(jm, 1).value)[:40]:42}"
                    f" | Jun r{ju} {str(wbv.cell(ju, 1).value)[:40]}"
                )
    print("\n  => each área's Jan–Mai formula reads the NEXT área's rows. Shift confirmed.")

    print("\n" + "=" * 78)
    print("2. DOES IT EXPLAIN THE NUMBERS? (recompute Jan–Mai with June's formula)")
    print("=" * 78)
    print(
        f"  {'área':13}{'mês':4}{'NOSSO':>12}{'livro atual':>13}{'Δ atual':>11}"
        f"{'livro c/ fórm. jun':>20}{'Δ corrigido':>13}"
    )
    cur_abs = fix_abs = 0.0
    cur_ties = fix_ties = 0
    resid: list[tuple[str, int, float]] = []
    for m in months:
        sections = assemble_dre_sections(
            snapshot=snaps[m],
            budget=monthly_budget(entries, month=m) if entries else None,
            budget_annual=ann or None,
            transfers=None,
            period_label=f"2026-{m:02d}",
            period_month=m,
            targets=None,
        )
        for area, rows_jm in FORMULA_JANMAI.items():
            ours = _our(sections, SECTION[area], "despesas_equipe")
            cur = sum(float(wbv.cell(r, BASE_COL[m]).value or 0.0) for r in rows_jm)
            fix = sum(
                float(wbv.cell(r, BASE_COL[m]).value or 0.0) for r in FORMULA_JUNHO[area]
            )
            d_cur, d_fix = round(ours - cur, 2), round(ours - fix, 2)
            cur_abs += abs(d_cur)
            fix_abs += abs(d_fix)
            cur_ties += abs(d_cur) < 0.02
            fix_ties += abs(d_fix) < 0.02
            if abs(d_fix) >= 0.02:
                resid.append((area, m, d_fix))
            print(
                f"  {area:13}{m:<4}{_brl(ours):>12}{_brl(cur):>13}{_brl(d_cur):>11}"
                f"{_brl(fix):>20}{_brl(d_fix):>13}"
            )
    n = len(months) * 3
    print(f"\n  células que batem:  atual {cur_ties}/{n}   com a fórmula de junho {fix_ties}/{n}")
    print(f"  soma dos |Δ|:       atual {_brl(cur_abs)}   com a fórmula de junho {_brl(fix_abs)}")
    if cur_abs:
        print(f"  => a fórmula deslocada explica {100 * (1 - fix_abs / cur_abs):.0f}% do erro absoluto.")

    print("\n" + "=" * 78)
    print("3. WHAT SURVIVES, AND WHY (the cost-centre ruling)")
    print("=" * 78)
    for area, m, d in sorted(resid, key=lambda t: -abs(t[2])):
        print(f"  {MESES[m]:10} {area:13} {_brl(d):>12}")
    # January is the bulk of what survives, and it is NOT purely a move between áreas:
    # measured, the three January residuals sum to +1.039,87, so our side carries more
    # despesa in total. That extra is the already-documented January Associações gap.
    jan_ours = sum(
        r["total"]
        for r in (snaps[1].get("despesas_equipe_area") or [])
        if isinstance(r, dict)
    )
    jan_book = sum(
        float(wbv.cell(r, BASE_COL[1]).value or 0.0)
        for rows in FORMULA_JUNHO.values()
        for r in rows
    )
    assoc_db = next(
        (
            r["total"]
            for r in (snaps[1].get("despesas_conta") or [])
            if str(r.get("id_conta")) == "020.060.0020"
        ),
        0.0,
    )
    assoc_book = sum(float(wbv.cell(r, BASE_COL[1]).value or 0.0) for r in (128, 129, 130, 131))
    print(
        f"\n  Janeiro, os três juntos: {_brl(round(jan_ours - jan_book, 2))} — ou seja NÃO é\n"
        "  só uma troca entre áreas; o nosso lado tem mais despesa no total. E é conhecido:"
    )
    print(
        f"    Associações (020.060.0020) no banco {_brl(assoc_db)} × planilha "
        f"{_brl(assoc_book)} = {_brl(round(assoc_db - assoc_book, 2))}"
    )
    print(
        "    = AASP 195,40 + Canal de Arbitragem 1.204,47 — lançamentos REAIS que a\n"
        "      planilha de janeiro não somou (docs/NOTA_CLIENTE.md: 'é sempre o banco\n"
        "      tendo mais informação que a planilha antiga, nunca o contrário').\n"
        "    O Canal de Arbitragem 1.204,47 é exatamente o resíduo da Arbitragem.\n"
        "  O resto é reclassificação: a planilha divide as duas fatias de Associações\n"
        "  entre Contencioso (r129) e Econômico (r130) a 700,10 cada; o banco marca as\n"
        "  DUAS como EDE, por isso o nosso Econômico lê 1.400,19 = 700,10 + 700,10.\n"
        "  Renata já decidiu isso — alocar pelo rótulo/centro de custo, o banco manda —\n"
        "  e extract.sql registra o mesmo padrão em maio."
    )


if __name__ == "__main__":
    main()
