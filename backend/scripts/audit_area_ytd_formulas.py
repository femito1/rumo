"""Audit the workbook's per-área YTD against ours, and explain every gap.

Answers TODO §5.2 of HANDOFF_2026-07-29: "Contencioso + Arbitragem differ ~7k,
Econômico ties". The client and the previous session both suspected a cross-área
leak on OUR side ("pode ser de lá, veio para cá"). It is a leak — but it is in the
**workbook's own Jan–May formulas**, not in our derivation.

What this script proves, all from `reference/workbook/Fechamento MBC 06.2026.xlsx`
plus the live snapshots:

1. The workbook's YTD cells are plain sums of its own monthly columns
   (`AB = C+G+K+O+S+W`), so any YTD gap decomposes exactly into per-month gaps.
   Nothing YTD-specific is broken in `accumulate_ytd`.

2. `Base_Resultado Mensal_V2` rows 204/205/206 ("Despesas Equipe" per área) point
   at DIFFERENT leaf rows in June than in Jan–May. The Despesas Área block is
   ordered Arbitragem / Contencioso / Econômico / Institucional within each expense
   family, and for five families (Eventos e Happy hour, Material Gráfico,
   Patrocínio, Refeições, Viagens) the Jan–May formulas are off by exactly one row
   — each área reads the label one line BELOW its own:

       Contencioso reads  '... - Direito Econômico'   (should be '... - Contencioso')
       Econômico   reads  '... - Institucional'       (should be '... - Direito Econômico')
       Arbitragem  reads  '... - Contencioso'         (should be '... - Arbitragem e Compliance')

   Net effect on the sheet: r141/145/149/153/157 (the *Institucional* rows) get
   counted, and r138/142/146/150/154 (the *Arbitragem* rows) are dropped entirely.
   That is a genuine one-row cross-área shift — exactly the shape Adriana guessed,
   just on her side of the wall. June's formulas were repaired and are correct,
   which is why June ties us to the centavo.

3. "Econômico ties" is an ARTIFACT of netting, not a sign that Econômico is right.
   Per-month it has the LARGEST gross error of the three áreas; ~94% of it cancels
   between over- and under-statements across Jan–May, so the YTD total looks clean.
   Contencioso cancels 74%, Arbitragem only 34% — which is why those two show a
   visible residue and Econômico does not. Do NOT read a matching YTD total as a
   validated área.

Because the pool `r207 "Despesa para ratear" = r198 − r203` is fed by r203 =
r204+r205+r206, the same off-by-one row also perturbs every área's Despesa
Institucional rateio in Jan–May.

⚠ Jan–May snapshots are ALSO still `extract_version` v1 (stale — pre-`f5fe22c`),
so their numbers move again after `backfill.ps1` re-runs on MBC-LDESK01. Re-run
this script after the backfill before taking any per-line Jan–Abr diff to Renata
(HANDOFF §4/§5.6): the workbook-formula finding below is independent of the
staleness, but the residual magnitudes are not.

Run: cd backend && python -m scripts.audit_area_ytd_formulas
"""
from __future__ import annotations

from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parents[2]
WORKBOOK = REPO / "reference" / "workbook" / "Fechamento MBC 06.2026.xlsx"

#: 'Areas Sintetico atualizado' Realizado column per month (Jan..Jun) and the YTD col.
#: That sheet blocks each month as Orçado|Realizado|Variação|Desvio, so Realizado
#: lands on 3, 7, 11, ... — do NOT reuse these to index Base_Resultado.
REAL_COL = {1: 3, 2: 7, 3: 11, 4: 15, 5: 19, 6: 23}
YTD_COL = 28

#: 'Base_Resultado Mensal_V2' month column (Jan..Jun) — one plain column per month,
#: C..H. A different grid from the sintetico sheet above.
BASE_COL = {1: 3, 2: 4, 3: 5, 4: 6, 5: 7, 6: 8}

#: 'Areas Sintetico atualizado' rows, per área.
SINTETICO_ROWS = {
    "contencioso": {"custo_equipe": 39, "despesas_equipe": 41, "despesa_institucional": 42},
    "economico": {"custo_equipe": 57, "despesas_equipe": 59, "despesa_institucional": 60},
    "arbitragem": {"custo_equipe": 75, "despesas_equipe": 77, "despesa_institucional": 78},
}

#: Base_Resultado leaf rows the "Despesas Equipe" per-área formulas add up.
#: The Jan–May sets are the BUGGY ones (shifted one row up for five families);
#: the June sets are correct — each leaf's own label names the área it belongs to.
LEAF_JAN_MAY = {
    "contencioso": [125, 129, 140, 144, 148, 152, 156, 160],
    "economico": [126, 130, 141, 145, 149, 153, 157, 161],
    "arbitragem": [127, 131, 139, 143, 147, 151, 155, 159],
}
LEAF_JUNE = {
    "contencioso": [125, 129, 139, 143, 147, 151, 155, 160],
    "economico": [126, 130, 140, 144, 148, 152, 156, 161],
    "arbitragem": [127, 131, 138, 142, 146, 150, 154, 159],
}


def _fmt(v: float) -> str:
    return f"{v:,.2f}"


def main() -> None:
    values = openpyxl.load_workbook(WORKBOOK, data_only=True)
    formulas = openpyxl.load_workbook(WORKBOOK, data_only=False)
    sint = values["Areas Sintetico atualizado"]
    base = values["Base_Resultado Mensal_V2"]
    base_f = formulas["Base_Resultado Mensal_V2"]

    print("=" * 100)
    print("1. The workbook's YTD is a plain sum of its own monthly columns")
    print("=" * 100)
    for area, rows in SINTETICO_ROWS.items():
        for key, row in rows.items():
            monthly = sum(sint.cell(row, c).value or 0 for c in REAL_COL.values())
            ytd = sint.cell(row, YTD_COL).value or 0
            flag = "OK" if abs(monthly - ytd) < 0.02 else "MISMATCH"
            print(f"  {area:12s} {key:22s} Σmonths {_fmt(monthly):>13s}  YTD {_fmt(ytd):>13s}  {flag}")

    print()
    print("=" * 100)
    print("2. r204/205/206 point at DIFFERENT leaf rows in June than in Jan–May")
    print("=" * 100)
    for row, area in ((204, "contencioso"), (205, "economico"), (206, "arbitragem")):
        print(f"\n  Base_Resultado r{row} ({area}) — formula per month:")
        for m, c in BASE_COL.items():
            print(f"     m{m}: {base_f.cell(row, c).value}")

    only_janmay = sorted(set(sum(LEAF_JAN_MAY.values(), [])) - set(sum(LEAF_JUNE.values(), [])))
    only_june = sorted(set(sum(LEAF_JUNE.values(), [])) - set(sum(LEAF_JAN_MAY.values(), [])))
    print("\n  Leaf rows Jan–May counts but June does NOT (these are INSTITUCIONAL rows):")
    for r in only_janmay:
        print(f"     r{r}: {base.cell(r, 1).value}")
    print("  Leaf rows June counts but Jan–May does NOT (these are ARBITRAGEM rows):")
    for r in only_june:
        print(f"     r{r}: {base.cell(r, 1).value}")

    print("\n  Consequence for r203 'Despesas Área' (and hence the r207 rateio pool):")
    print(f"     {'m':>3s}{'Arb OMITTED':>16s}{'Inst WRONGLY in':>20s}{'net r203 error':>18s}")
    for m in range(1, 6):
        c = BASE_COL[m]
        omitted = sum(base.cell(r, c).value or 0 for r in only_june)
        wrong = sum(base.cell(r, c).value or 0 for r in only_janmay)
        print(f"     {m:>3d}{_fmt(omitted):>16s}{_fmt(wrong):>20s}{_fmt(wrong - omitted):>18s}")

    print()
    print("=" * 100)
    print("3. Re-mapping Jan–May onto June's (correct) rows moves the workbook, not us")
    print("=" * 100)
    print(f"  {'area':13s}{'book as shipped':>18s}{'book re-mapped':>18s}{'shift':>14s}")
    for area in LEAF_JAN_MAY:
        shipped = sum(
            sum(base.cell(r, BASE_COL[m]).value or 0 for r in LEAF_JAN_MAY[area])
            for m in range(1, 6)
        ) + sum(base.cell(r, BASE_COL[6]).value or 0 for r in LEAF_JUNE[area])
        remapped = sum(
            sum(base.cell(r, BASE_COL[m]).value or 0 for r in LEAF_JUNE[area])
            for m in range(1, 7)
        )
        print(f"  {area:13s}{_fmt(shipped):>18s}{_fmt(remapped):>18s}{_fmt(remapped - shipped):>14s}")

    print()
    print("  → Take the per-line Jan–Abr diff to Renata as a WORKBOOK formula question")
    print("    (rows 204/205/206, five families off by one row), not as a DB fix.")
    print("    Re-run after backfill.ps1: Jan–May snapshots are still extract v1.")


if __name__ == "__main__":
    main()
