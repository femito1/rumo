"""Regenerate the May 3-way comparison spreadsheet (client-facing proof-of-match).

3 legs, per institucional line + per-area block:
  * Planilha  — raw cell in reference/workbook/Fechamento MBC 05.2026.xlsx ("Areas Sintetico atualizado", May col 19)
  * Alvo      — the extracted/overridden target from workbook_targets_2026.json (targets_for("2026-05"))
  * Sistema   — assemble_dre_sections(...) on the real May snapshot (targets=None -> raw DB numbers)

Output: reference/comparativo/Comparativo_MBC_Maio_2026.xlsx (color-coded).
Companion narrative: docs/NOTA_MAIO_2026.md.

⚠ SNAPSHOT SOURCE: the local test fixture (tests/fixtures/sisjuri_2026_05.json) is STALE —
it predates the líquido/desdobramento/prof blocks, so it produces a false +2.854 despesas
gap and a wrong per-area recebimento split. This script therefore rebuilds the May snapshot
from the values captured off the live box on 2026-07-21 (the `_MAY_OVERRIDES` below), which
reproduce the sacred number (415.927,84) and the live-prod figures (RL 29.691,61 / reserva
2.969,16). To refresh from a new snapshot, replace _MAY_OVERRIDES with the fresh
closing_2026-05.json blocks (despesas_liquido, despesas_desdobramento, vale_adm,
despesas_equipe_area, recebimento_area_prof, comissao_deriv, custo_equipe_area).

Run: cd backend && python -m scripts.build_may_comparison
"""
from __future__ import annotations

import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.closing.dre import assemble_dre_sections
from app.closing.workbook_targets import targets_for

REPO = Path(__file__).resolve().parents[2]
FIXTURE = REPO / "backend" / "tests" / "fixtures" / "sisjuri_2026_05.json"
WORKBOOK = REPO / "reference" / "workbook" / "Fechamento MBC 05.2026.xlsx"
OUT = REPO / "reference" / "comparativo" / "Comparativo_MBC_Maio_2026.xlsx"

# Fresh May blocks captured from the live box (closing_2026-05.json, 2026-07-21). These
# OVERRIDE the stale fixture's equivalents so the assembler reproduces live-prod numbers.
_MAY_OVERRIDES = {
    "despesas_liquido": [
        {"id_conta": "020.010.0010", "liquido": 27477.67, "bruto": 27477.67, "n": 1},
        {"id_conta": "020.010.0020", "liquido": 4996, "bruto": 4996, "n": 1},
        {"id_conta": "020.010.0030", "liquido": 863.59, "bruto": 863.59, "n": 1},
        {"id_conta": "020.010.0040", "liquido": 6916.97, "bruto": 6916.97, "n": 1},
        {"id_conta": "020.020.0010", "liquido": 65.29, "bruto": 65.29, "n": 1},
        {"id_conta": "020.020.0030", "liquido": 759.72, "bruto": 759.72, "n": 3},
        {"id_conta": "020.030.0020", "liquido": 630, "bruto": 630, "n": 1},
        {"id_conta": "020.030.0080", "liquido": 21.94, "bruto": 21.94, "n": 1},
        {"id_conta": "020.030.0100", "liquido": 1914.41, "bruto": 1966.76, "n": 2},
        {"id_conta": "020.030.0140", "liquido": 55.6, "bruto": 55.6, "n": 1},
        {"id_conta": "020.030.0150", "liquido": 215, "bruto": 215, "n": 1},
        {"id_conta": "020.040.0010", "liquido": 9252.45, "bruto": 9653.09, "n": 3},
        {"id_conta": "020.040.0030", "liquido": 3346.68, "bruto": 3984.15, "n": 1},
        {"id_conta": "020.040.0050", "liquido": 8042.94, "bruto": 8570, "n": 1},
        {"id_conta": "020.050.0010", "liquido": 4498.5, "bruto": 4498.5, "n": 1},
        {"id_conta": "020.050.0020", "liquido": 3250.01, "bruto": 3250.01, "n": 1},
        {"id_conta": "020.050.0060", "liquido": 400, "bruto": 400, "n": 1},
        {"id_conta": "020.060.0020", "liquido": 1204.47, "bruto": 1204.47, "n": 1},
        {"id_conta": "020.060.0040", "liquido": 182.71, "bruto": 182.71, "n": 1},
        {"id_conta": "020.090.0010", "liquido": 1426.72, "bruto": 1426.72, "n": 2},
        {"id_conta": "030.010.0180", "liquido": 1600, "bruto": 1600, "n": 1},
        {"id_conta": "040.030.0010", "liquido": 14705.8, "bruto": 14705.8, "n": 1},
        {"id_conta": "040.040.0030", "liquido": 7129.1, "bruto": 7341.98, "n": 5},
    ],
    "despesas_desdobramento": [
        {"id_conta": "020.050.0050", "valor": 1791.5, "historico": "GPS INSS"},
        {"id_conta": "020.050.0050", "valor": 25, "historico": "GPS RAT"},
        {"id_conta": "020.090.0040", "valor": 59.98, "historico": "bolo"},
        {"id_conta": "020.090.0040", "valor": 146, "historico": "pao WM"},
        {"id_conta": "040.040.0010", "valor": 904.68, "historico": "notebooks"},
        {"id_conta": "040.040.0020", "valor": 399.99, "historico": "SSD"},
        {"id_conta": "020.030.0020", "valor": 257.36, "historico": "flanelas"},
        {"id_conta": "020.030.0020", "valor": 2166.53, "historico": "Claude software"},
        {"id_conta": "020.060.0020", "valor": 700.1, "historico": "IBRAC"},
        {"id_conta": "020.060.0020", "valor": 700.09, "historico": "IBRAC"},
        {"id_conta": "020.060.0020", "valor": 217.4, "historico": "AASP"},
        {"id_conta": "040.040.0030", "valor": 110, "historico": "Adobe"},
        {"id_conta": "020.050.0110", "valor": 1269.46, "historico": "Convenio mla"},
        {"id_conta": "020.030.0060", "valor": 968.1, "historico": "Vale"},
    ],
    "vale_adm": 3326.94,
    "custo_equipe_area": [
        {"sigla": "JVO", "id_conta": "030.010.0100/0220", "valor": 1236.9},
        {"sigla": "VSR", "id_conta": "030.010.0100/0220", "valor": 75.6},
    ],
    "despesas_equipe_area": [
        {"cc": "ECT", "total": 917.49, "n": 2},
        {"cc": "EDE", "total": 3804.82, "n": 4},
        {"cc": "ESP", "total": 1272.47, "n": 2},
    ],
    "recebimento_area_prof": [
        {"grupo": "Administração", "total": -0.01, "fat": -0.01},
        {"grupo": "Arbitragem", "total": 41997.5, "fat": 219430.24},
        {"grupo": "Equipe Ambiental", "total": -138.15, "fat": 5.97},
        {"grupo": "Equipe Contencioso", "total": 240444.72, "fat": 336677.36},
        {"grupo": "Equipe Direito Econômico", "total": 166875.57, "fat": 177649.16},
        {"grupo": "NãoAlocados", "total": -33251.8, "fat": -13774.67},
    ],
    "comissao_deriv": [{"kind": "lawyer", "sigla": "EHF", "area": None, "valor": 2128.06}],
}

INST = [
    ("recebimento", 4, "Recebimento (receita de honorários)"),
    ("custo_equipe", 6, "Custos Diretos (custo equipe + comissão)"),
    ("despesas", 13, "Despesas Institucionais"),
    ("resultado_bruto", 25, "Resultado Bruto"),
    ("imposto", 28, "Imposto (15% do recebimento)"),
    ("amortizacao", 29, "Amortização"),
    ("resultado_liquido", 30, "Resultado Líquido"),
    ("reserva_bonus", 32, "Reserva de Bônus"),
]
AREAS = [("contencioso", 35, "Contencioso"), ("economico", 53, "Econômico"), ("arbitragem", 71, "Arbitragem")]
ALINES = [("recebimento", 1, "Recebimento"), ("custo_equipe", 4, "Custo equipe"), ("resultado_bruto", 8, "Resultado Bruto")]
MAY_COL = 19


def fresh_may() -> dict:
    snap = json.loads(FIXTURE.read_text(encoding="utf-8"))
    snap.update(_MAY_OVERRIDES)
    return snap


def _explain(line: str, raw, d, scope: str) -> tuple[str, str]:
    if raw is None or d is None:
        return ("—", "")
    if abs(d - raw) <= 1.0:
        return ("bate exato", "Sistema reproduz a planilha (diferença ≤ R$1 = arredondamento).")
    if scope == "inst":
        if line == "despesas":
            return ("sistema mais correto", "Aluguel líquido da sublocação Belline: banco 24.359,77 vs planilha 24.230,60 (+129,17). Renata: usar o banco.")
        if line in ("resultado_bruto", "resultado_liquido"):
            return ("consequência do aluguel", "Reflete o ±129,17 do aluguel.")
        if line == "reserva_bonus":
            return ("consequência do aluguel", "10% do líquido corrigido.")
    else:
        if line == "resultado_bruto":
            return ("regrupamento (Renata)", "Despesas Área alocadas pelo rótulo/centro de custo (fórmula da planilha tinha deslocamento de 1 linha em Viagens). Soma das 3 áreas é a mesma.")
    return ("verificar", "")


def main() -> None:
    snap = fresh_may()
    db = assemble_dre_sections(snapshot=snap, budget={}, period_label="2026-05", period_month=5, targets=None)
    tg = targets_for("2026-05") or {}

    def dbval(key: str, line: str):
        for row in db.get(key, {}).get("rows", []):
            if row.get("key") == line:
                r = row.get("Realizado")
                return r.get("value") if isinstance(r, dict) else r
        return None

    ws_raw = openpyxl.load_workbook(WORKBOOK, data_only=True)["Areas Sintetico atualizado"]

    def rawcell(r: int):
        v = ws_raw.cell(r, MAY_COL).value
        return round(float(v), 2) if isinstance(v, (int, float)) else None

    wb = openpyxl.Workbook()
    sh = wb.active
    sh.title = "Comparativo Maio 2026"
    hdr_fill = PatternFill("solid", fgColor="1F3864")
    sec_fill = PatternFill("solid", fgColor="D9E1F2")
    green = PatternFill("solid", fgColor="C6EFCE")
    amber = PatternFill("solid", fgColor="FFEB9C")
    grey = PatternFill("solid", fgColor="E7E6E6")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(wrap_text=True, vertical="top")
    rightal = Alignment(horizontal="right")

    cols = ["Linha", "Planilha (célula)", "Alvo extraído", "Sistema (banco)", "Diferença (sis−plan)", "Categoria", "Explicação"]
    for i, (c, w) in enumerate(zip(cols, [38, 16, 16, 16, 17, 22, 60]), 1):
        cell = sh.cell(1, i, c)
        cell.fill = hdr_fill
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.border = border
        cell.alignment = wrap
        sh.column_dimensions[get_column_letter(i)].width = w

    r = [2]

    def section(title: str) -> None:
        for j in range(1, 8):
            c = sh.cell(r[0], j, title if j == 1 else None)
            c.fill = sec_fill
            c.font = Font(bold=True, size=11)
            c.border = border
        r[0] += 1

    def put(label: str, raw, t, d, scope: str) -> None:
        dif = round(d - raw, 2) if (isinstance(d, (int, float)) and isinstance(raw, (int, float))) else None
        cat, exp = _explain(scope_line[0], raw, d, scope)
        for j, val in enumerate([label, raw, t, d, dif, cat, exp], 1):
            c = sh.cell(r[0], j, val)
            c.border = border
            c.font = Font(size=10)
            if j in (2, 3, 4, 5) and isinstance(val, (int, float)):
                c.number_format = "#,##0.00"
                c.alignment = rightal
            if j == 7:
                c.alignment = wrap
        fill = green if cat == "bate exato" else (amber if "correto" in cat else grey)
        for j in range(1, 8):
            sh.cell(r[0], j).fill = fill
        r[0] += 1

    scope_line = [""]
    section("INSTITUCIONAL")
    for key, row, label in INST:
        scope_line[0] = key
        put(label, rawcell(row), tg.get("institucional", {}).get(key), dbval("institucional", key), "inst")
    for area, base, aname in AREAS:
        section(f"ÁREA — {aname}")
        for key, off, label in ALINES:
            scope_line[0] = key
            put(label, rawcell(base + off), tg.get(area, {}).get(key), dbval(area, key), "area")

    r[0] += 1
    for note in (
        "Legenda: verde = bate exato · amarelo = banco mais completo/correto · cinza = consequência/arredondamento.",
        "Recebimento por área usa a base do Demonstrativo por profissional (DB_RESULTADO_PROF), a mesma do LegalDesk.",
    ):
        sh.cell(r[0], 1, note).font = Font(italic=True, size=9)
        r[0] += 1

    sh.freeze_panes = "A2"
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"saved {OUT}")
    # sanity: sacred + live figures
    assert abs(dbval("institucional", "recebimento") - 415927.84) < 0.01
    assert abs(dbval("institucional", "resultado_liquido") - 29691.61) < 0.01
    print("sanity OK: recebimento 415.927,84 / RL 29.691,61")


if __name__ == "__main__":
    main()
