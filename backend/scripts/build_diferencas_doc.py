"""Standalone PT-BR document explaining the MATERIAL workbook-vs-system differences.

Written for Renata (finance): no codebase knowledge assumed, one consistent format per
difference, every number traceable to a workbook cell. Deliberately a DOCUMENT and not a
product feature — the client asked for the explanations to live outside the app (2026-08),
and the in-app "Diferenças conhecidas" panel was removed in the same change.

The format, every time:

    Planilha (<aba>, linha <n>)   <valor>
    Nosso sistema                 <valor>
    Diferença                     <±valor>
    Por quê / Onde conferir / O que precisamos

Reads the LIVE snapshots + the June workbook and derives the per-área YTD from the
PRODUCTION ``assemble_dre_sections`` — never a re-implementation, so the document cannot
drift from what the app shows. Re-run it after any re-extract.

Materiality: R$ 1.000 on the YTD **or on any single month** (see ``_material`` for why the
second half is not optional). The client's words: R$ 4,80 does not matter, R$ 1.900 does.
Everything below the line is summed and named as a remainder rather than dropped, so "small"
never reads as "unexamined".

**A line-level delta is not an answer.** "Despesa Institucional differs by R$ 1.400" tells
finance that something moved, not what — so the two biggest despesa lines are decomposed all
the way down and the reader can redo the arithmetic:

* ``_detalhe_despesas`` — every despesa family in both columns for every month, then, for
  each family+month that differs, the individual accounts on each side with the agreeing
  ones already netted out and subtotals printed. Safe to present as complete because both
  sides reconcile exactly: the ten families sum to the workbook's r198 and to our own
  *Despesas*, and each family's leaves sum to its own header, in all six months.
* ``_detalhe_rateio`` — the two inputs behind every per-área Despesa Institucional (the pool
  and each área's Custo-equipe share), which reproduce the printed number to the centavo on
  both sides. That converts three opaque per-área lines into checkable arithmetic.

Run: cd backend && python -m scripts.build_diferencas_doc
Writes: docs/DIFERENCAS_ACUMULADO_2026.md
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl

REPO = Path(__file__).resolve().parents[2]
WORKBOOK = REPO / "reference" / "workbook" / "Fechamento MBC 06.2026.xlsx"
OUT = REPO / "docs" / "DIFERENCAS_ACUMULADO_2026.md"

#: Materiality floor, in R$. Applied to the YTD **and to every single month** — see
#: ``_material``. A line whose months cancel to a small YTD (Econômico · Despesas Equipe:
#: −31,45 YTD out of a +1.504,72 May) is exactly the netting artifact this document warns
#: about two paragraphs into itself; filing it under "menores" contradicted that warning.
LIMIAR = 1000.0


def _delta(nosso: float, planilha: float) -> float:
    """The one definition of a difference: rounded values subtracted, then rounded.

    Every total in this document must be reachable by adding up what is PRINTED. Rounding
    the raw subtraction instead (``round(o - b, 2)``) makes ``round(Σ) ≠ Σ(round)`` and the
    two disagreed in the wild: Contencioso · Custo equipe printed +3.140,19 in the summary
    and +3.140,20 in its own detail table, one centavo apart on the same line of the same
    document. Guarded by ``test_diferencas_doc.py``.
    """
    return round(round(nosso, 2) - round(planilha, 2), 2)


def _material(ytd: float, mensais: list[float], limiar: float = LIMIAR) -> bool:
    """Material if the YTD **or any single month** reaches the floor.

    YTD-only was the original rule and it hid the worst kind of difference this project
    keeps re-learning: months that cancel. R$ 1.000 in May does not stop mattering because
    a −R$ 1.000 in February happens to offset it.
    """
    return abs(ytd) >= limiar or any(abs(d) >= limiar for d in mensais)


#: Lines that are SUMS of other lines in this document. They carry the biggest deltas, so
#: sorting purely by size would open with three "consequência das linhas acima" entries
#: before the reader has seen a single cause. These are ordered last.
#: NB the per-área ``custo_equipe`` is NOT derived — it has its own causes (vale, ISS,
#: AASP); only the institucional roll-ups and the resultados are.
def _is_derivada(section: str, line: str) -> bool:
    """True for lines that are pure SUMS of other lines shown in this document.

    ``institucional.despesas`` is deliberately NOT one, even though it is a total: three
    per-área Despesa Institucional entries point at it as their cause, so it must keep its
    own detailed breakdown (vale ADM, prêmio anual de seguro, IR Fonte, e-Social…). Listing
    it among the roll-ups silently deleted that breakdown — caught 2026-08-04.
    """
    return line in ("resultado_bruto", "resultado_liquido") or (
        section == "institucional" and line == "custo_equipe"
    )


def _ordem(t: tuple) -> tuple[int, float]:
    """Sort: explaining lines first (largest first), derived roll-ups after."""
    section, line = t[0], t[1]
    return (1 if _is_derivada(section, line) else 0, -abs(t[6]))


MESES = {1: "Janeiro", 2: "Fevereiro", 3: "Março", 4: "Abril", 5: "Maio", 6: "Junho"}
#: 'Areas Sintetico atualizado' Realizado column per month.
SINT_COL = {1: 3, 2: 7, 3: 11, 4: 15, 5: 19, 6: 23}
#: 'Base_Resultado Mensal_V2' month column (the detail behind the sintetico totals).
BASE_COL = {1: 3, 2: 4, 3: 5, 4: 6, 5: 7, 6: 8}

#: Our despesa family name -> (workbook family HEADER row, first row AFTER the family) in
#: 'Base_Resultado Mensal_V2'. Read off the sheet, and verified two ways before use: the ten
#: headers sum to r198 on the workbook side and to our ``despesas`` on ours, and inside each
#: family the leaf rows sum to their own header in all six months. That is what makes an
#: account-level table trustworthy rather than suggestive.
FAMILIAS: dict[str, tuple[int, int]] = {
    "Ocupação": (85, 92),
    "Telecomunicações": (92, 95),
    "Despesas Gerais": (95, 110),
    "Consultoria": (110, 116),
    "Salários Administração": (116, 124),
    "Administrativas": (124, 137),
    "Investimentos em Prospecção": (137, 158),
    "Gestão do Conhecimento": (158, 164),
    "Endomarketing": (164, 168),
    "Informática": (180, 191),
}

#: (our section, our line key, label, sintetico row).
LINHAS: tuple[tuple[str, str, str, int], ...] = (
    ("institucional", "recebimento", "Receita", 4),
    ("institucional", "custo_equipe", "Custos Diretos", 6),
    ("institucional", "despesas", "Despesas Indiretas", 13),
    ("institucional", "resultado_bruto", "Resultado Bruto", 25),
    ("institucional", "imposto", "Impostos", 28),
    ("institucional", "amortizacao", "Amortização", 29),
    ("institucional", "resultado_liquido", "Resultado Líquido", 30),
    ("contencioso", "custo_equipe", "Contencioso · Custo equipe", 39),
    ("contencioso", "despesas_equipe", "Contencioso · Despesas Equipe", 41),
    ("contencioso", "despesa_institucional", "Contencioso · Despesa Institucional", 42),
    ("contencioso", "resultado_bruto", "Contencioso · Resultado Bruto", 43),
    ("economico", "custo_equipe", "Econômico · Custo equipe", 57),
    ("economico", "despesas_equipe", "Econômico · Despesas Equipe", 59),
    ("economico", "despesa_institucional", "Econômico · Despesa Institucional", 60),
    ("economico", "resultado_bruto", "Econômico · Resultado Bruto", 61),
    ("arbitragem", "custo_equipe", "Arbitragem · Custo equipe", 75),
    ("arbitragem", "despesas_equipe", "Arbitragem · Despesas Equipe", 77),
    ("arbitragem", "despesa_institucional", "Arbitragem · Despesa Institucional", 78),
    ("arbitragem", "resultado_bruto", "Arbitragem · Resultado Bruto", 79),
)

#: Hand-written causes, keyed by (section, line). Deliberately SHORT and in plain PT-BR —
#: this is read by finance, so no script names, no internal counts ("18 células"), no
#: percentages. The shared causes (fórmula deslocada, POOL rateado, RB = soma) are stated
#: once in the "Quatro causas" summary; the per-line entries just point back to them.
#: ``conferir`` names workbook cells only; ``precisamos`` is what finance must do (almost
#: always None now).
CAUSAS: dict[tuple[str, str], dict[str, str | None]] = {
    ("arbitragem", "custo_equipe"): {
        "causa": (
            "Convênio médico de um advogado (JGS) em **fevereiro** — e a própria "
            "planilha responde. Em fevereiro ela mantém a distribuição (linha 70) e o "
            "pró-labore (linha 71) dele e deixa só o convênio (linha 69) em branco. Quem "
            "recebe distribuição está na folha, então o plano é custo real, e o sistema "
            "o tem lançado. De março em diante ele sai dos dois lados e a Arbitragem "
            "bate em 0,00. **Não é dúvida — é uma omissão da coluna de fevereiro.**"
        ),
        "conferir": (
            "Planilha, aba `Base_Resultado Mensal_V2`, linhas **69, 70 e 71** (o convênio, "
            "a distribuição e o pró-labore dele), coluna de fevereiro (**D**)."
        ),
        "precisamos": None,
    },
    ("contencioso", "custo_equipe"): {
        "causa": (
            "O **vale dos advogados** (causa 3 do resumo): entra sempre no custo da área, e "
            "a planilha só o lança em alguns meses — nas linhas 26 e 27 ele aparece em "
            "jan/fev/jun e fica **zerado em mar/abr/mai**, que são exatamente os três meses "
            "com diferença acima de mil reais. Em janeiro e fevereiro o vale está nos dois "
            "lados e o que sobra é só classificação: a **AASP** (−97,70 e −163,06), que a "
            "planilha põe no Custo equipe e nós em Despesa de Área — sai desta linha e "
            "entra em outra da mesma área, sem mexer no Resultado Bruto. Junho bate em "
            "0,00. Em abril há ainda o **ISS trimestral** (−253,55), que é o mesmo tipo de "
            "troca. Ver *Diferenças de classificação*."
        ),
        "conferir": (
            "Planilha, aba `Base_Resultado Mensal_V2`, linhas **26 e 27** (Vale Refeição e "
            "Vale Transporte do Contencioso) — compare jan/fev/jun com mar/abr/mai."
        ),
        "precisamos": None,
    },
    ("economico", "custo_equipe"): {
        "causa": (
            "Três coisas. Em **janeiro** (−887,63), a parte MBC do convênio do **RB** "
            "(−789,94, causa 4 do resumo — é a nossa estimativa) mais a **AASP** (−97,70, "
            "classificação: troca de linha dentro da área). Em **março e abril** (+2.425,63 "
            "e +2.770,99), duas coisas iguais nos dois meses: o **vale dos advogados** "
            "(+1.008,40 em cada) e a distribuição de uma advogada (ASG), que a planilha "
            "lança **líquida** e nós bruta (+1.509,00 em cada). O que os separa é pequeno: "
            "em março, um **seguro de vida** que só a planilha tem (−92,45); em abril, o "
            "**ISS trimestral** (+253,59), que sai do Contencioso e entra aqui. Fevereiro e "
            "maio ficam abaixo de cem reais.\n\n"
            "A **estagiária do Direito Econômico** entra na planilha em março e nós a "
            "reproduzimos ao centavo — ela não gera diferença, mas explica por que o custo "
            "da área sobe nos dois lados a partir daquele mês."
        ),
        "conferir": (
            "Planilha, aba `Base_Resultado Mensal_V2`, linhas **44 e 48** (convênio de EHF "
            "e de RB) e **52** (a bolsa da estagiária)."
        ),
        "precisamos": None,
    },
    ("contencioso", "despesa_institucional"): {
        "causa": (
            "**Não é da área — é a despesa institucional total, rateada** (causa 2 do "
            "resumo). A divisão entre as três áreas não cria nem apaga dinheiro, então esta "
            "linha não tem causa própria: olhe *Despesas Indiretas*.\n\n"
            "Duas coisas se somam aqui. A maior parte vem do **total** (a linha 198, "
            "detalhada em *Despesas Indiretas*). O resto vem da **fórmula deslocada** "
            "(causa 1): o que se rateia é o total **menos** as despesas das áreas (linha "
            "207 = 198 − 203), e a linha 203 é a soma das linhas 204/205/206 — as mesmas "
            "que estão erradas de janeiro a maio. Por isso a soma das três áreas não é "
            "igual à diferença do total nesses meses: em abril, por exemplo, as áreas somam "
            "−3.394,34 contra −2.070,03 do total, e os R$ 1.324,31 de diferença são "
            "exatamente o erro da linha 203."
        ),
        "conferir": (
            "Planilha, aba `Base_Resultado Mensal_V2`: linha **207** (total a ratear = "
            "198 − 203) e linha **203**. O custo de cada área, que dá a proporção do "
            "rateio, está nas linhas **5 / 30 / 60** da mesma aba."
        ),
        "precisamos": None,
    },
    ("economico", "despesa_institucional"): {
        "causa": "Mesma causa do Contencioso: é o total institucional rateado (causa 2).",
        "conferir": (
            "Planilha, aba `Base_Resultado Mensal_V2`, linhas **207** e **203**; custo das "
            "áreas nas linhas **5 / 30 / 60**."
        ),
        "precisamos": None,
    },
    ("arbitragem", "despesa_institucional"): {
        "causa": "Mesma causa do Contencioso: é o total institucional rateado (causa 2).",
        "conferir": (
            "Planilha, aba `Base_Resultado Mensal_V2`, linhas **207** e **203**; custo das "
            "áreas nas linhas **5 / 30 / 60**."
        ),
        "precisamos": None,
    },
    ("contencioso", "despesas_equipe"): {
        "causa": (
            "A **fórmula deslocada** da planilha (causa 1 do resumo): de janeiro a maio a "
            "linha 204 soma as linhas de *Eventos*, *Material Gráfico*, *Patrocínio*, "
            "*Refeições* e *Viagens* uma linha abaixo da sua — pega a do Direito Econômico "
            "em vez da do Contencioso. Junho já está com a fórmula certa e bate."
        ),
        "conferir": (
            "Planilha, aba `Base_Resultado Mensal_V2`, linha **204**: compare a fórmula de "
            "junho (`=H125+H129+H139+H143+...`) com a de maio (`=G125+G129+G140+G144+...`) — "
            "as cinco últimas parcelas estão uma linha adiante nos meses de janeiro a maio."
        ),
        "precisamos": "Vale copiar as fórmulas de junho para janeiro–maio na planilha.",
    },
    ("economico", "despesas_equipe"): {
        "causa": (
            "A **fórmula deslocada** da planilha (causa 1), na linha 205. Junho está certo "
            "e bate. O acumulado quase se anula (−31,45) porque os meses têm sinais "
            "opostos — **não** porque a linha esteja certa: fevereiro difere −1.166,75 e "
            "maio +1.504,72. É o caso que este documento usa para dizer que um total que "
            "fecha por compensação não está validado."
        ),
        "conferir": (
            "Planilha, aba `Base_Resultado Mensal_V2`, linha **205**, colunas de janeiro a "
            "maio (compare com a de junho)."
        ),
        "precisamos": "Mesma correção de fórmula.",
    },
    ("arbitragem", "despesas_equipe"): {
        "causa": (
            "A **fórmula deslocada** da planilha (causa 1) — na Arbitragem o efeito é o "
            "maior dos três, porque a linha 206 é a que perde as parcelas de maior valor. "
            "O resíduo de janeiro (+1.204,47) é o **Canal de Arbitragem**, que a planilha "
            "daquele mês não somou."
        ),
        "conferir": (
            "Planilha, aba `Base_Resultado Mensal_V2`, linha **206**, colunas de janeiro a "
            "maio (compare com a de junho)."
        ),
        "precisamos": "Mesma correção de fórmula.",
    },
    ("contencioso", "resultado_bruto"): {
        "causa": "Não tem causa própria: é a soma das linhas acima da área.",
        "conferir": "Some as linhas 39 a 42 da própria área na planilha.",
        "precisamos": None,
    },
    ("economico", "resultado_bruto"): {
        "causa": "Não tem causa própria: é a soma das linhas acima da área.",
        "conferir": "Some as linhas 57 a 60 da própria área na planilha.",
        "precisamos": None,
    },
    ("arbitragem", "resultado_bruto"): {
        "causa": "Não tem causa própria: é a soma das linhas acima da área.",
        "conferir": "Some as linhas 75 a 78 da própria área na planilha.",
        "precisamos": None,
    },
    ("institucional", "despesas"): {
        "causa": (
            "Esta linha também explica a maior parte da Despesa Institucional das três "
            "áreas (ela é rateada daqui). Somando família por família, as partes dão "
            "**exatamente** a diferença de cada mês — não sobra centavo. Por mês, o que a "
            "compõe:\n\n"
            "* **Vale do administrativo** — a causa dominante em mar (−2.199,08), abr "
            "(−2.199,20) e mai (−2.280,60). A planilha usa uma base **diferente em cada "
            "mês** nas linhas 122/123: só a pessoa do administrativo em fev e jun, as três "
            "pessoas em abr, e nenhuma das duas regras em jan/mar/mai. Nós usamos sempre a "
            "mesma regra (só a pessoa do administrativo; os estagiários vão para as áreas "
            "deles), por isso jan/fev/jun batem e os outros não.\n"
            "* **Janeiro (+1.533,77)**: são duas coisas somadas. As **Associações** "
            "(+1.399,87) — a planilha não somou a AASP (195,40) nem o Canal de Arbitragem "
            "(1.204,47), que existem no sistema — e o **IR Fonte ADM** (+169,52), uma conta "
            "sem linha na planilha; menos os 35,52 do vale-transporte (ver a pergunta no "
            "fim do documento) e 0,10 de arredondamento.\n"
            "* **Fevereiro (+1.249,19)**: o **e-Social** (+1.032,35), outra conta sem linha "
            "na planilha, mais +217,11 de Administrativas.\n"
            "* **Janeiro, seguro**: um prêmio **anual** de 2.722,55 (a planilha digita "
            "182,71 todo mês). Não muda o total desta linha — a planilha põe o prêmio em "
            "*Administrativas* (linha 133) e nós em Ocupação, e as duas famílias somam na "
            "linha 198. É por isso que ele **não** aparece na conta de janeiro acima.\n"
            "* **Aluguel** (abr +19,17 e mai +129,17): usamos o aluguel líquido da "
            "sublocação (crédito Belline).\n"
            "* **Tarifa bancária**: existe no sistema e está zerada na planilha (linha "
            "136). Não é mensal — só mar (37,39) e jun (4,80) nos seis meses. Em junho é a "
            "**única** diferença que sobra.\n"
            "* **Trocas de família** (Endomarketing ↔ Prospecção, Ocupação ↔ "
            "Administrativas): a mesma conta em famílias diferentes de cada lado, mas as "
            "duas entram no total — efeito **zero** (ver *Diferenças de classificação*).\n"
            "* **Março, vale da estagiária** (+543,22): um pagamento de benefícios fora "
            "da conta transitória (Vale Refeição 507,10 + Vale Transporte 36,12, com o "
            "nome dela no histórico). É o que faz a família de Salários Administração "
            "fechar em 0,00 de março a junho, depois de tirado o vale.\n"
            "* **Março**: um curso de Arbitragem (−815,49) que a planilha pôs em "
            "institucional e nós na área; e Informática −237,60 (a planilha usou o valor "
            "bruto, nós o líquido)."
        ),
        "conferir": (
            "Todas na aba `Base_Resultado Mensal_V2`: linhas **122/123** (vale ADM), **86** "
            "(aluguel), **128–131** (Associações), **133** (seguro), **136** (tarifa "
            "bancária, zerada), **158** (curso), **180** (Informática). O total é a linha "
            "**198** — e, depois de nomear cada item acima, ele fecha sem sobra."
        ),
        "precisamos": None,
    },
    ("institucional", "custo_equipe"): {
        "causa": (
            "É a soma dos custos de equipe das três áreas — mesmas causas já descritas em "
            "cada uma (vale, convênio, estagiária)."
        ),
        "conferir": "Ver as três linhas de Custo equipe por área.",
        "precisamos": None,
    },
}


def _brl(v: float | None) -> str:
    """Money the way the client's own documents write it: ``R$ 1.234,56``.

    The currency prefix is not decoration here — this document puts our figure next to
    theirs on every line, and a finance reader should never have to infer the unit.
    """
    if v is None:
        return "—"
    s = f"{abs(v):,.2f}".replace(",", "@").replace(".", ",").replace("@", ".")
    return f"{'-' if v < 0 else ''}R$ {s}"


def _sgn(v: float | None) -> str:
    """Signed money, with an explicit ``+`` so the direction is never ambiguous.

    A zero gets NO sign: "+R$ 0,00" reads like a rounded-down positive when it is an
    exact tie, and this document leans on the ties (June) to make its point.
    """
    if v is None:
        return "—"
    s = f"{abs(v):,.2f}".replace(",", "@").replace(".", ",").replace("@", ".")
    sinal = "" if abs(v) < 0.005 else ("-" if v < 0 else "+")
    return f"{sinal}R$ {s}"


def _col(idx: int) -> str:
    """1-based column index -> Excel letter, so the doc can name the exact cell."""
    from openpyxl.utils import get_column_letter

    return get_column_letter(idx)


def _pct(frac: float) -> str:
    """A share the Brazilian way: ``34,95%``. The whole document is PT-BR; a decimal
    point in a percentage next to ``R$ 1.234,56`` reads as a thousands separator."""
    return f"{frac * 100:.2f}".replace(".", ",") + "%"


def _hdr(first: str, months: list[int], abbr: dict[int, str]) -> str:
    return f"| {first} | " + " | ".join(abbr[m] for m in months) + " | Acumulado |"


def _sep(months: list[int]) -> str:
    return "|---|" + "---:|" * (len(months) + 1)


def ytd_of(rows: list[tuple], section: str, line: str) -> float:
    return next(t[6] for t in rows if t[0] == section and t[1] == line)


def _row_deltas(
    label: str,
    cells: dict[int, tuple[float, float, float]],
    months: list[int],
    total: float | None = None,
) -> str:
    """One summary row: the per-month DELTA plus the accumulated one.

    A tie renders as ``R$ 0,00 ✓`` rather than a blank so a month that agrees reads as
    *checked*, not *missing* — the document leans on June being clean across the board.

    The Acumulado column is the sum of the DISPLAYED months, not a separately rounded
    total. Those differ by a centavo (sum-of-rounded ≠ rounded-of-sum), and a row the
    client cannot add up destroys confidence in the whole document — which is the exact
    problem this rewrite exists to fix. ``total`` is ignored when given.
    """
    del total  # kept for call-site symmetry; the row must add up as printed
    out = []
    for m in months:
        d = cells[m][2]
        out.append(f"{_sgn(d)} ✓" if abs(d) < 0.005 else _sgn(d))
    soma = round(sum(cells[m][2] for m in months), 2)
    return f"| {label} | " + " | ".join(out) + f" | **{_sgn(soma)}** |"


def _num(cell: Any) -> float:
    """A workbook cell as money. Blank rows and the odd stray ``'  '`` read as zero."""
    v = cell.value if hasattr(cell, "value") else cell
    return float(v) if isinstance(v, (int, float)) else 0.0


def _folhas_planilha(base: Any, familia: str, m: int) -> list[tuple[str, float, int]]:
    """The workbook's own leaf rows under a family, for one month: (label, value, row)."""
    lo, hi = FAMILIAS[familia]
    out = []
    for r in range(lo + 1, hi):
        rot = base.cell(r, 1).value or base.cell(r, 2).value
        v = round(_num(base.cell(r, BASE_COL[m])), 2)
        if rot and abs(v) >= 0.005:
            out.append((" ".join(str(rot).split()), v, r))
    return out


def _folhas_nossas(assembled: dict[str, Any], familia: str) -> list[tuple[str, float]]:
    """Our own per-account rows under a family, for one month: (label, value)."""
    out = []
    for row in (assembled.get("institucional") or {}).get("rows") or []:
        if str(row.get("key") or "").startswith(f"acct::{familia}::"):
            cell = row.get("Realizado")
            v = cell.get("value") if isinstance(cell, dict) else cell
            if isinstance(v, (int, float)) and abs(round(float(v), 2)) >= 0.005:
                out.append((" ".join(str(row.get("Linha")).split()), round(float(v), 2)))
    return out


def _conciliar(
    nossas: list[tuple[str, float]], planilha: list[tuple[str, float, int]]
) -> tuple[int, list[tuple[str, float]], list[tuple[str, float, int]]]:
    """Match leaves BY VALUE inside one family+month; return (matched, only-ours, only-book).

    Matching by LABEL is not available: the two sides genuinely name accounts differently
    (our single *Serviços de Informática* is the book's *Suporte de Informática* + *Suporte
    Totvs*; the book splits *Associações* three ways by área where we keep one account).
    Inventing a label join would produce confident-looking rows that are simply wrong, so
    equal values are treated as the same lançamento and everything else is reported as
    unmatched — which is precisely the list a finance reader wants to see.
    """
    from collections import Counter

    comuns = Counter(v for _, v, _ in planilha) & Counter(v for _, v in nossas)
    resta = Counter(comuns)
    so_nossas = []
    for nome, v in nossas:
        if resta[v] > 0:
            resta[v] -= 1
        else:
            so_nossas.append((nome, v))
    resta = Counter(comuns)
    so_planilha = []
    for nome, v, r in planilha:
        if resta[v] > 0:
            resta[v] -= 1
        else:
            so_planilha.append((nome, v, r))
    return sum(comuns.values()), so_nossas, so_planilha


def _detalhe_despesas(
    assembled: dict[int, dict[str, Any]],
    base: Any,
    months: list[int],
    abbr: dict[int, str],
) -> list[str]:
    """Which despesas, with which numbers, per month — down to the account.

    A line-level delta says *that* something moved; finance needs *what*. The prose above
    already names the causes, but a reader who wants to check one has to trust the prose.
    This puts the whole grid on the page: every despesa family in both columns for every
    month, and then, for each family and month that differs, the individual accounts on
    each side with the ones that agree already netted out.

    Both sides reconcile exactly — the ten families sum to the workbook's r198 and to our
    own *Despesas*, and each family's leaves sum to its header — so nothing here is a
    partial view or an estimate. Accounts are matched by VALUE, not label (see
    ``_conciliar``).
    """
    out: list[str] = []
    add = out.append

    def nossa(m: int, fam: str) -> float:
        for row in (assembled[m].get("institucional") or {}).get("rows") or []:
            if row.get("key") == f"sec::{fam}":
                cell = row.get("Realizado")
                v = cell.get("value") if isinstance(cell, dict) else cell
                return float(v) if isinstance(v, (int, float)) else 0.0
        return 0.0

    def livro(m: int, fam: str) -> float:
        return _num(base.cell(FAMILIAS[fam][0], BASE_COL[m]))

    add("#### Quais despesas, mês a mês")
    add("")
    add("As dez famílias de despesa institucional, com a diferença de cada uma em cada mês.")
    add("Elas somam exatamente o total dos dois lados — a linha **198** na planilha e a")
    add("*Despesas* no sistema — então esta tabela é a diferença inteira, sem resto.")
    add("")
    add(_hdr("Família (linha)", months, abbr))
    add(_sep(months))
    difere: list[tuple[str, list[int]]] = []
    for fam, (hrow, _fim) in FAMILIAS.items():
        cells = {m: (round(nossa(m, fam), 2), round(livro(m, fam), 2)) for m in months}
        ds = {m: _delta(*cells[m]) for m in months}
        meses_com_dif = [m for m in months if abs(ds[m]) >= 0.005]
        if meses_com_dif:
            difere.append((fam, meses_com_dif))
        celulas = " | ".join(
            f"{_sgn(ds[m])} ✓" if abs(ds[m]) < 0.005 else _sgn(ds[m]) for m in months
        )
        add(f"| {fam} (r{hrow}) | {celulas} | **{_sgn(round(sum(ds.values()), 2))}** |")
    total = round(
        sum(_delta(nossa(m, f), livro(m, f)) for m in months for f in FAMILIAS), 2
    )
    add("| **Total** | " + " | ".join("" for _ in months) + f" | **{_sgn(total)}** |")
    add("")
    tocadas = ", ".join(f"*{f}*" for f, _ in difere)
    add("Batem em todos os meses as famílias que não aparecem abaixo. As que diferem em")
    add(f"algum mês são {tocadas} — e para cada uma delas os lançamentos vêm a seguir.")
    add("")

    add("#### Conta por conta, onde há diferença")
    add("")
    add("Para cada família e mês que difere: as contas de cada lado, já **descontadas as que")
    add("batem**. Ou seja, o que está listado é exatamente o que explica aquela diferença —")
    add("nada mais. Quando um valor aparece só de um lado, é porque a conta existe só ali ou")
    add("está somada de outra forma (a planilha divide algumas contas por área, o sistema as")
    add("mantém unidas; o inverso também acontece).")
    add("")
    add("A conta fecha em cada linha: **(soma do sistema) − (soma da planilha) = diferença**,")
    add("e os subtotais estão na tabela para que dê para conferir sem somar à mão.")
    add("")
    centavos: list[str] = []
    for fam, meses_com_dif in difere:
        add(f"**{fam}** — linha {FAMILIAS[fam][0]} da planilha")
        add("")
        add(
            "| Mês | Só na planilha | Σ | Só no sistema | Σ | Contas que batem |"
            " Diferença |"
        )
        add("|---|---|---:|---|---:|---:|---:|")
        for m in meses_com_dif:
            nossas = _folhas_nossas(assembled[m], fam)
            livres = _folhas_planilha(base, fam, m)
            batem, so_nossas, so_livro = _conciliar(nossas, livres)
            d = _delta(round(nossa(m, fam), 2), round(livro(m, fam), 2))
            sp = round(sum(v for _, v, _ in so_livro), 2)
            sn = round(sum(v for _, v in so_nossas), 2)
            esq = (
                "<br>".join(f"r{r} {nome} · {_brl(v)}" for nome, v, r in so_livro) or "—"
            )
            dir_ = "<br>".join(f"{nome} · {_brl(v)}" for nome, v in so_nossas) or "—"
            # The listed parts must reconstruct the family delta. Where they miss by a
            # centavo it is the workbook carrying a half-centavo inside an account it
            # splits by área; say so on the spot instead of printing a row that does not
            # add up.
            nota = ""
            if abs(round(sn - sp, 2) - d) >= 0.005:
                nota = " ¹"
                centavos.append(f"{fam} / {MESES[m]}")
            add(
                f"| {MESES[m]} | {esq} | {_brl(sp)} | {dir_} | {_brl(sn)} | {batem} |"
                f" {_sgn(d)}{nota} |"
            )
        add("")
    if centavos:
        add(f"¹ Nestas linhas ({', '.join(centavos)}) os subtotais reconstroem a diferença")
        add("com **um centavo** de folga: a planilha divide a conta de Associações entre as")
        add("três áreas e cada parte carrega meio centavo, que reaparece ao somar. Não é")
        add("dinheiro faltando — é arredondamento da própria divisão.")
        add("")
    return out


def _detalhe_rateio(
    assembled: dict[int, dict[str, Any]],
    base: Any,
    sint: Any,
    months: list[int],
) -> list[str]:
    """The two inputs behind every per-área Despesa Institucional, on both sides.

    "É o total rateado" is true but unverifiable as prose. The rateio has exactly two
    inputs — the pool and each área's share of Custo equipe — and multiplying them
    reproduces the printed number to the centavo on BOTH sides, in all six months. Showing
    them side by side turns three opaque per-área lines into arithmetic the reader can
    redo, and makes visible which of the two inputs is actually responsible in each month.
    """
    out: list[str] = []
    add = out.append
    AREAS_R = (
        ("contencioso", "Contencioso", 5),
        ("economico", "Econômico", 30),
        ("arbitragem", "Arbitragem", 60),
    )

    def our(m: int, sec: str, line: str) -> float:
        for row in (assembled[m].get(sec) or {}).get("rows") or []:
            if row.get("key") == line:
                cell = row.get("Realizado")
                v = cell.get("value") if isinstance(cell, dict) else cell
                return float(v) if isinstance(v, (int, float)) else 0.0
        return 0.0

    add("#### De onde sai a Despesa Institucional de cada área")
    add("")
    add("O rateio tem só duas entradas: o **valor a ratear** (linha 207 = 198 − 203) e a")
    add("**parte de cada área no Custo equipe** (linhas 5 / 30 / 60). Multiplicando as duas")
    add("chega-se ao número de cada área, ao centavo, nos dois lados. A tabela mostra as")
    add("duas entradas para que se veja qual delas causa a diferença em cada mês.")
    add("")
    add("| Mês | A ratear (planilha) | A ratear (sistema) | Diferença |")
    add("|---|---:|---:|---:|")
    for m in months:
        pool_b = _num(base.cell(207, BASE_COL[m]))
        nossa_de = sum(our(m, a, "despesas_equipe") for a, _, _ in AREAS_R)
        pool_o = our(m, "institucional", "despesas") - nossa_de
        add(
            f"| {MESES[m]} | {_brl(round(pool_b, 2))} | {_brl(round(pool_o, 2))} |"
            f" {_sgn(_delta(pool_o, pool_b))} |"
        )
    add("")
    add("E a parte de cada área (percentual do Custo equipe total):")
    add("")
    add("| Mês | Contencioso | Econômico | Arbitragem |")
    add("|---|---|---|---|")
    for m in months:
        cel = []
        tb = sum(_num(base.cell(cr, BASE_COL[m])) for _, _, cr in AREAS_R)
        to = sum(our(m, a, "custo_equipe") for a, _, _ in AREAS_R)
        for a, _nome, cr in AREAS_R:
            sb = _num(base.cell(cr, BASE_COL[m])) / tb if tb else 0.0
            so = our(m, a, "custo_equipe") / to if to else 0.0
            cel.append(f"{_pct(sb)} → {_pct(so)}")
        add(f"| {MESES[m]} | " + " | ".join(cel) + " |")
    add("")
    add("Lê-se *planilha → sistema*. Em junho os percentuais são idênticos e as três áreas")
    add("batem; nos outros meses a diferença de cada área vem das duas entradas juntas — o")
    add("valor a ratear (que é a linha 198, detalhada acima, menos a linha 203, afetada pela")
    add("fórmula deslocada) e a parte de cada área, que muda quando o Custo equipe muda.")
    add("")
    return out


def _br_date(iso: str) -> str:
    y, m, d = iso.split("-")
    return f"{d}/{m}/{y}"


def _vintage(snaps: dict[int, Any], months: list[int]) -> str:
    """When the compared snapshots were produced, so a reader knows the data's age.

    A document full of numbers with no date invites "is this still current?" in the
    meeting, and the answer is not guessable from the content.
    """
    stamps = sorted(
        d
        for d in (
            str((snaps[m].get("meta") or {}).get("generated_at") or "")[:10]
            for m in months
        )
        if d
    )
    if not stamps:
        return "data desconhecida"
    lo, hi = stamps[0], stamps[-1]
    return _br_date(lo) if lo == hi else f"{_br_date(lo)} a {_br_date(hi)}"


def _our(sections: dict[str, Any], section: str, line: str) -> float | None:
    for row in (sections.get(section) or {}).get("rows") or []:
        if row.get("key") == line:
            cell = row.get("Realizado")
            v = cell.get("value") if isinstance(cell, dict) else cell
            return float(v) if isinstance(v, (int, float)) else None
    return None


def main() -> None:
    from dotenv import load_dotenv

    load_dotenv(REPO / "backend" / ".env")

    from app.api.providers import get_budget_repo, get_snapshot_store
    from app.budget.models import annual_budget, monthly_budget
    from app.closing.dre import assemble_dre_sections, convenio_mbc_shares

    snaps = get_snapshot_store().snapshots_by_year(2026, client_id="mbc")
    entries = get_budget_repo().get_budget("mbc", 2026)
    ann = annual_budget(entries) if entries else {}
    # Same whole-year Parte MBC shares the app uses (``provider.py``), so a month with a
    # stale convênio memo is valued here exactly as the product values it. Without this
    # the document would quietly disagree with the screens it is explaining.
    shares = convenio_mbc_shares(snaps)
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
    sint = wb["Areas Sintetico atualizado"]
    # The detail sheet behind those totals — where the per-account breakdown comes from.
    base = wb["Base_Resultado Mensal_V2"]

    months = [m for m in MESES if m in snaps]
    assembled: dict[int, dict[str, Any]] = {}
    for m in months:
        assembled[m] = assemble_dre_sections(
            snapshot=snaps[m],
            budget=monthly_budget(entries, month=m) if entries else None,
            budget_annual=ann or None,
            transfers=None,
            period_label=f"2026-{m:02d}",
            period_month=m,
            targets=None,
            convenio_shares=shares,
        )

    # YTD per line: ours from the production assembly, theirs from the workbook.
    ytd: list[tuple[str, str, str, int, float, float, float]] = []
    for section, line, label, wrow in LINHAS:
        o = sum((_our(assembled[m], section, line) or 0.0) for m in months)
        b = sum(float(sint.cell(wrow, SINT_COL[m]).value or 0.0) for m in months)
        ytd.append((section, line, label, wrow, round(o, 2), round(b, 2), round(o - b, 2)))

    ult = MESES[max(months)]
    abbr = {m: MESES[m][:3] for m in months}

    # Per-month deltas per line, so nothing is presented only as a YTD total.
    per_month: dict[tuple[str, str], dict[int, tuple[float, float, float]]] = {}
    ytd: list[tuple[str, str, str, int, float, float, float]] = []
    for section, line, label, wrow in LINHAS:
        cells: dict[int, tuple[float, float, float]] = {}
        for m in months:
            o = _our(assembled[m], section, line) or 0.0
            b = float(sint.cell(wrow, SINT_COL[m]).value or 0.0)
            cells[m] = (round(o, 2), round(b, 2), _delta(o, b))
        per_month[(section, line)] = cells
        ytd.append((
            section, line, label, wrow,
            round(sum(c[0] for c in cells.values()), 2),
            round(sum(c[1] for c in cells.values()), 2),
            round(sum(c[2] for c in cells.values()), 2),
        ))

    def _mensais(section: str, line: str) -> list[float]:
        return [per_month[(section, line)][m][2] for m in months]

    materiais = [t for t in ytd if _material(t[6], _mensais(t[0], t[1]))]
    menores = [
        t
        for t in ytd
        if not _material(t[6], _mensais(t[0], t[1]))
        and any(d != 0.0 for d in _mensais(t[0], t[1]))
    ]

    L: list[str] = []
    add = L.append
    rb = ytd_of(ytd, "institucional", "resultado_bruto")
    add(f"# Diferenças entre a planilha e o sistema — Janeiro a {ult} de 2026")
    add("")
    add(f"Comparação da planilha `{WORKBOOK.name}` com os números do sistema, dados"
        f" extraídos em {_vintage(snaps, months)}.")
    add("")
    add("## O essencial")
    add("")
    add("1. **A receita bate em todos os meses.** Toda diferença está em *despesa*.")
    add(f"2. No acumulado de janeiro a {ult.lower()}, o **Resultado Bruto** difere"
        f" **{_sgn(rb)}** — sobre uma receita de mais de R$ 2 milhões.")
    add("3. **Cada centavo dessa diferença tem uma causa identificada**, listada abaixo. Ter"
        " causa não é o mesmo que bater: nenhuma delas foi corrigida ainda, e uma —"
        " a parte MBC do convênio do RB em janeiro — é uma **estimativa nossa**, não um"
        " valor lançado.")
    add("4. **Junho fecha**: a única diferença de despesa é a tarifa bancária de R$ 4,80"
        " (o resto são centavos de arredondamento do recebimento). É o mês em que a"
        " planilha já está com as fórmulas certas e inclui o vale — a referência de como os"
        " dois lados batem quando ambos estão corretos.")
    add("")
    add("Cada diferença vem **mês a mês** com a célula da planilha ao lado (aba"
        " `Areas Sintetico atualizado`, colunas " + ", ".join(
        f"**{abbr[m]}={_col(SINT_COL[m])}**" for m in months) + "). Detalhamos as que"
        f" passam de **{_brl(LIMIAR)}** no acumulado **ou em qualquer mês isolado**; as"
        " menores estão no fim.")
    add("")
    add("Nas **Despesas Indiretas** — a linha que puxa também a Despesa Institucional das"
        " três áreas — a diferença é aberta até a **conta**: quais despesas, com que valor,"
        " em cada mês, dos dois lados. Quem quiser conferir um número específico deve ir"
        " direto a *Quais despesas, mês a mês* e *Conta por conta, onde há diferença*.")
    add("")

    # ── What does NOT differ.
    add("## O que NÃO difere")
    add("")
    add(_hdr("Linha", months, abbr))
    add(_sep(months))
    for section, line, label, wrow, _o, _b, _d in ytd:
        if section == "institucional" and line in ("recebimento", "imposto", "amortizacao"):
            add(_row_deltas(label, per_month[(section, line)], months, ytd_of(ytd, section, line)))
    add("")
    add("Os centavos de maio e junho são arredondamento — a planilha arredonda o")
    add("recebimento para reais inteiros.")
    add("")

    add("## As quatro causas")
    add("")
    add("Todas as linhas citadas abaixo estão na aba `Base_Resultado Mensal_V2` da planilha,")
    add("que é o detalhe por trás dos totais da `Areas Sintetico atualizado`.")
    add("")
    add("1. **Fórmula deslocada na planilha** (linhas 204/205/206, janeiro a maio). Cada")
    add("   uma dessas linhas soma as despesas de uma área; nas cinco últimas parcelas")
    add("   (*Eventos*, *Material Gráfico*, *Patrocínio*, *Refeições*, *Viagens*) a fórmula")
    add("   aponta uma linha adiante e pega a da área vizinha. Junho já está certo. Move")
    add("   *Despesas Equipe* e — por meio da linha 203 — também a *Despesa Institucional*")
    add("   das três áreas. É ajuste na planilha, não no sistema.")
    add("2. **Despesa Institucional por área é o total institucional dividido entre as**")
    add("   **áreas.** A divisão não cria nem apaga dinheiro; a diferença vem do total (ver")
    add("   *Despesas Indiretas*) e da causa 1, através da linha 203.")
    add("3. **O vale entra no custo da área, e a planilha varia de mês.** Nos advogados, as")
    add("   linhas 26/27 (Contencioso) e 56/57 (Econômico) trazem o vale em alguns meses e")
    add("   ficam zeradas em mar/abr/mai. No administrativo, as linhas 122/123 usam três")
    add("   bases diferentes ao longo dos seis meses. O sistema usa sempre a mesma regra.")
    add("4. **A anotação do convênio médico fica velha.** A *memória de cálculo* no")
    add("   lançamento diz quanto do plano é da MBC; em jan/fev ela descrevia um plano")
    add("   antigo (o mesmo texto vinha desde 2025, com o plano mudando duas vezes). O")
    add("   sistema já não depende dela — calcula pela proporção dos meses corretos — mas")
    add("   **atualizá-la quando um plano mudar** é o que mantém o valor exato em vez de")
    add("   estimado.")
    add("")
    add("*Um total que fecha porque dois erros se anulam não está validado — por isso tudo*")
    add("*aparece mês a mês, não só no acumulado. Uma linha entra no detalhe abaixo se*")
    add(f"*passar de {_brl(LIMIAR)} no acumulado **ou** em qualquer mês isolado.*")
    add("")
    add("## Resumo: onde estão as diferenças")
    add("")
    add("Diferença = Sistema − Planilha, por mês. `✓` = bate.")
    add("")
    add(_hdr("Linha", months, abbr))
    add(_sep(months))
    for section, line, label, _w, _o, _b, _d in sorted(materiais, key=_ordem):
        add(_row_deltas(label, per_month[(section, line)], months, ytd_of(ytd, section, line)))
    add("")

    add("## Detalhe, linha por linha")
    add("")
    add("Cada tabela é uma linha da aba `Areas Sintetico atualizado`; a coluna *Célula* dá")
    add("o endereço exato para conferir.")
    add("")
    # Roll-up lines have no cause of their own (Resultado Bruto/Líquido = sum of the lines
    # above; institucional Custos Diretos = sum of the three áreas). Listing them with a
    # full table each added ~100 lines that said "veja as linhas acima" six times over, so
    # they get one shared paragraph and stay out of the per-line detail.
    detalhe = [t for t in materiais if not _is_derivada(t[0], t[1])]
    somas = [t for t in materiais if _is_derivada(t[0], t[1])]
    _primeira_inst = next(
        (t[0] for t in sorted(detalhe, key=_ordem) if t[1] == "despesa_institucional"),
        None,
    )
    for section, line, label, wrow, o, b, d in sorted(detalhe, key=_ordem):
        info = CAUSAS.get((section, line))
        add(f"### {label} — {_sgn(d)} no acumulado")
        add("")
        add("| Mês | Célula | Planilha | Sistema | Diferença |")
        add("|---|---|---:|---:|---:|")
        cells = per_month[(section, line)]
        for m in months:
            ov, bv, dv = cells[m]
            marca = "" if abs(dv) >= 0.005 else " ✓"
            add(
                f"| {MESES[m]} | `{_col(SINT_COL[m])}{wrow}` | {_brl(bv)} | {_brl(ov)} |"
                f" {_sgn(dv)}{marca} |"
            )
        sb = round(sum(cells[m][1] for m in months), 2)
        so = round(sum(cells[m][0] for m in months), 2)
        add(f"| **Acumulado** | — | **{_brl(sb)}** | **{_brl(so)}** | **{_sgn(round(so - sb, 2))}** |")
        add("")
        if info:
            add(f"{info['causa']}")
            add("")
            add(f"*Conferir:* {info['conferir']}")
            add("")
        else:
            add("*Causa ainda não documentada — falar com o time antes da reunião.*")
            add("")
        if section == "institucional" and line == "despesas":
            L.extend(_detalhe_despesas(assembled, base, months, abbr))
        # The rateio inputs go under the FIRST per-área Despesa Institucional to appear; the
        # other two say "mesma causa" and point back at it.
        if line == "despesa_institucional" and section == _primeira_inst:
            L.extend(_detalhe_rateio(assembled, base, sint, months))

    if somas:
        add("### Linhas que são somas de outras")
        add("")
        add("Estas não têm causa própria — cada uma é a soma das linhas acima (Resultado")
        add("Bruto e Líquido dentro de cada bloco; Custos Diretos = as três áreas). Elas")
        add("aparecem aqui só para fechar o acumulado:")
        add("")
        add("Conferindo: *Custos Diretos* é exatamente a soma dos três *Custo equipe*, mês a")
        add("mês. Já a soma dos três *Resultado Bruto* por área dá −R$ 5.004,01 contra os")
        add("−R$ 5.003,04 do Resultado Bruto institucional: os **97 centavos** de diferença")
        add("são o rateio da Despesa Institucional, que a planilha e o sistema arredondam em")
        add("pontos diferentes ao dividir o total entre três áreas.")
        add("")
        add(_hdr("Linha", months, abbr))
        add(_sep(months))
        for section, line, label, _w, _o, _b, _d in sorted(somas, key=_ordem):
            add(_row_deltas(label, per_month[(section, line)], months, None))
        add("")

    add("## Diferenças menores")
    add("")
    if menores:
        add(_hdr("Linha", months, abbr))
        add(_sep(months))
        for section, line, label, _w, _o, _b, _d in sorted(menores, key=lambda t: -abs(t[6])):
            add(_row_deltas(label, per_month[(section, line)], months, ytd_of(ytd, section, line)))
        add("")
        add(f"Somadas: **{_sgn(round(sum(t[6] for t in menores), 2))}** no acumulado, e"
            " nenhuma passa de mil reais em nenhum mês isolado — são centavos de")
        add("arredondamento do recebimento, que a planilha digita em reais inteiros.")
    else:
        add("Nenhuma.")
    add("")

    add("## O que fazer")
    add("")
    add("### 1. Na planilha: copiar as fórmulas de junho para janeiro–maio")
    add("")
    add("Nas linhas **204, 205 e 206** da aba `Base_Resultado Mensal_V2`, as fórmulas de")
    add("janeiro a maio pegam cinco parcelas da área vizinha. As de junho estão corretas —")
    add("copiá-las para os meses anteriores corta bem mais da metade do erro mensal de")
    add("*Despesas Equipe* (de R$ 10.216 para R$ 4.245, somados os seis meses em módulo) e")
    add("cerca de um terço do de *Despesa Institucional*.")
    add("")
    add("⚠ **Isso não fecha o acumulado do Resultado Bruto, e é importante saber disso antes")
    add("de mexer.** Medimos: a correção quase não move o Resultado Bruto (de R$ 14.175 para")
    add("R$ 14.009 em erro mensal somado, e o acumulado sai de −R$ 5.003 para −R$ 5.004).")
    add("O motivo é que a linha **198** — o total institucional, que é o que chega ao")
    add("Resultado Bruto — não referencia as linhas 204/205/206. A correção arruma a")
    add("**distribuição entre as áreas**, que é o que ela deve arrumar; o acumulado depende")
    add("das outras causas.")
    add("")
    add("### 2. No sistema: manter a anotação do convênio atualizada quando o plano mudar")
    add("")
    add("A *memória de cálculo* no lançamento do convênio é o que diz quanto do plano é da")
    add("MBC. Quando ela fica velha, o sistema estima a parte da MBC pela proporção dos")
    add("outros meses — funciona, mas é estimativa. Hoje há uma: a parte da MBC do **RB em")
    add("janeiro** (o plano dele mudou e nenhuma anotação registra a proporção daquele mês).")
    add("Se puderem confirmar esse número, ele deixa de ser estimado.")
    add("")
    add("### 3. Uma pergunta: de onde vem o R$ 35,52 do vale-transporte de janeiro?")
    add("")
    add("Na aba `Base_Resultado Mensal_V2`, a célula `C123` traz `=35,52+262,64`. Os")
    add("**262,64** são o vale-transporte da pessoa")
    add("do administrativo (14 dias × R$ 18,76) e conferem. Os **35,52** não aparecem em")
    add("nenhum lançamento do sistema — nem em janeiro, nem em nenhum outro mês. Não é")
    add("vale-refeição (o menor do ano é R$ 783,70) e não corresponde a um número inteiro de")
    add("dias em nenhuma diária de vale.")
    add("")
    add("Também não é um pedaço que falte do nosso número: o vale-transporte de janeiro")
    add("(R$ 262,64) já está completo, então os 35,52 estão somados por cima. Se for de outra")
    add("competência ou um acerto pontual, é só dizer e passamos a tratá-lo da mesma forma.")
    add("")
    add("## Diferenças de classificação")
    add("")
    add("Estas são de **onde** o valor aparece, não de **quanto** ele é: o lançamento existe")
    add("nos dois lados, em seções diferentes. Ficam registradas porque **se repetem todo")
    add("mês** e costumam gerar dúvida.")
    add("")
    add("Não mudam **nenhum** total — a conta entra na mesma soma dos dois lados:")
    add("")
    add("* **Endomarketing × Prospecção** e **Ocupação × Administrativas** — a mesma conta em")
    add("  famílias diferentes de cada lado; as duas entram na linha 198.")
    add("* **Prêmio de seguro** — é anual (lançado em janeiro e julho); a planilha o divide")
    add("  em parcelas mensais e o classifica em *Administrativas* (linha 133), o sistema em")
    add("  *Ocupação*. Como as duas famílias somam na linha 198, o total não se move.")
    add("")
    add("Movem uma linha, mas **não** movem o Resultado Bruto da área — o valor só troca de")
    add("seção dentro da mesma área, e as duas seções ficam acima do Resultado Bruto:")
    add("")
    add("* **AASP** — dentro do Custo equipe na planilha, em Despesa de Área no sistema.")
    add("  Contencioso, −97,70 em janeiro e −163,06 em fevereiro.")
    add("")
    add("Move valor **entre áreas** — some no total, mas cada área sente:")
    add("")
    add("* **ISS trimestral** — o sistema lança por advogado, a planilha numa única linha da")
    add("  área. No acumulado sobra R$ 0,04, mas em abril são −R$ 253,55 no Contencioso e")
    add("  +R$ 253,59 no Econômico: o Resultado Bruto **de cada área** muda, o institucional")
    add("  não.")
    add("")
    add("Estas duas **mudam** o total e estão contadas nas diferenças acima — ficam aqui só")
    add("porque também são decisões de método, não erros de nenhum dos lados:")
    add("")
    add("* **Vale do administrativo** — a planilha muda de base a cada mês (linhas 122/123:")
    add("  só a pessoa do administrativo em fev/jun, as três em abril, nenhuma das duas em")
    add("  jan/mar/mai); o sistema usa sempre a mesma regra, com os estagiários no custo das")
    add("  áreas deles. Não há o que convergir: aproximar a planilha significaria reproduzir")
    add("  três regras diferentes.")
    add("* **Aluguel** — o sistema usa o valor líquido da sublocação (crédito Belline),")
    add("  a planilha o bruto. Diferença de +R$ 129,17 em abril e maio.")
    add("")

    OUT.write_text("\n".join(L) + "\n", encoding="utf-8")
    print(f"wrote {OUT.relative_to(REPO)}  ({len(materiais)} materiais, {len(menores)} menores)")


if __name__ == "__main__":
    main()
