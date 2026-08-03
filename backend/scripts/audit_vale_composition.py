"""What the workbook's Vale-ADM rows are actually made of, month by month.

Written to settle two hand-typed cells that had been open questions for weeks:
``Base_Resultado`` **C123** = ``=35,52+262,64`` (January) and **E123** = ``=543,22+674``
(March). The standing theory was "the first term is a vale-refeição and the second a
vale-transporte". Testing it against the data resolved most of the block and narrowed the
open part to two numbers.

WHAT THIS ESTABLISHED
---------------------
1. **The second terms are vale-TRANSPORTE, and they are identifiable exactly.**
   * January's ``262,64`` = MLA's own VT. The raw extrato spells the arithmetic out:
     *"Vale transporte / Calculo: 14 dias x R$ 18,76"*.
   * March's ``674`` = the **whole month's** VT for all three people
     (86,40 + 268,80 + 318,92 = 674,12). The June extrato states it verbatim:
     *"Vale Transporte: 22 dias = Total: 674,12"*.
   So the two months are NOT the same shape: January pairs its top-up with MLA-only,
   March pairs one with all three. That asymmetry is why a single explanation never fit.

2. **The first terms are NOT vale-refeição.** Every VR in the year is 46,10/day and lands
   between 783,70 and 1.014,20 (17–22 days). ``35,52`` is two orders of magnitude below the
   smallest possible monthly VR, so it cannot be one.

3. **Vale rates are per-person and per-day, and they are IN the histórico** — VSR 10,80,
   MLA 18,76, JVO 33,60, VR 46,10 for everyone. Every posted vale in the year is a whole
   number of days at one of those rates. **Neither 35,52 nor 543,22 is** — checked against
   every rate × 1..25 days and every pair of rates.

4. **April and May needed no top-up at all — they are simply the un-adjusted months.**
   ``655,36`` = 300,16 + 268,80 + 86,40 and ``607,04`` = 262,64 + 268,80 + 75,60: the full
   three-person VT, exactly the Vale-ADM issue already documented. February and June are
   MLA-only and tie our rule to the centavo.

543,22 IS SOLVED — AND THE "VR + VT" THEORY IS EXACTLY RIGHT FOR IT
------------------------------------------------------------------
March's ``543,22`` is a **separate estagiária payable, outside the transitória**, and it is
literally a vale-refeição plus a vale-transporte:

    020.080.0050  Vale Refeição    507,10   "Vale Refeição - Vitoria resende Março"
    020.080.0060  Vale Transporte   36,12   "Vale transporte - VRS - março"
                                   -------
                                    543,22

⚠ My first exhaustive sweep reported this as "not found anywhere". That was **WRONG**, and
the reason is worth remembering: 543,22 is not *stored* — it is the SUM of two stored rows,
and my broad pass only tested values and pairwise *differences*, never sums. It was also
already written down in ``docs/SISJURI_DB.md``. **Search for compositions, not just values,
and grep the durable docs before declaring something unknown.**

THE DAY-COUNT FRAMEWORK (the reusable result)
---------------------------------------------
Because every vale is a whole number of days, each month can be read as a pair of DAY
COUNTS, and that turns out to be a strong consistency check:

    mês   VR dias   VT dias (MLA)   gap
    Jan     18          14          +4
    Fev     22          18          +4
    Mar     20          17          +3
    Abr     20          16          +4
    Mai     17          14          +3
    Jun     22          17          +5

The gap is **+3 to +5 in every single month** — VR is paid for every worked day, VT only
for days actually commuted. January's +4 with VT at 14 days is dead centre of that pattern.

WHAT IS STILL OPEN: 35,52 ONLY — AND WE NOW KNOW WHAT IT IS *NOT*
-----------------------------------------------------------------
The most attractive hypothesis was a **typo**: 2 extra days of MLA's VT is
``2 × 18,76 = 37,52``, and ``262,64 + 37,52 = 300,16`` is exactly 16 days — a clean number,
one digit away from the 35,52 in the cell.

**The day-count table REFUTES it.** At 16 VT days January's gap would be **+2**, which
occurs in no other month; at 14 days it is +4, matching every other month. So MLA really
did have 14 commuting days in January, our ``262,64`` is her complete and correct VT, and
the 35,52 is something added ON TOP of a already-correct figure — not a missing piece of it.

That is the useful conclusion: **we are not missing a vale**. Everything else was checked
and excluded —

* not a whole number of days at any rate (46,10 / 10,80 / 18,76 / 33,60), nor a sum of two;
* absent from all 2026 snapshots **and all twelve 2025 months** (the year exists in the
  store — I had not looked before);
* absent from the raw May "Extrato de Contas" (.xls via xlrd, full untruncated histórico),
  the June extrato PDF, the Jan–Mai PPTX and the AR demonstrativo;
* not in any Excel cell comment or threaded comment — the workbook has 14 of them and they
  do annotate neighbouring rows ("IBRAC", "aasp", "E-CPF", "Aluguel - valor pago Belline"),
  but **none on r122/r123**;
* the same ``=35.52+262.64`` formula appears in the Feb workbook (as C118) and the May one,
  so it is a stable entry, not a slip in one copy;
* the ``020.080.*`` estagiário accounts that explain March do not exist in January at all.

⚠ Note for whoever re-probes this: ``despesas_desdobramento.historico`` is truncated to
**80 chars** by the extract (``SUBSTR(d.DESCHISTORICO,1,80)``) — which cuts off exactly
where the "Calculo: N dias x R$ X" text lives. The full text is only in the raw extrato
export. Widening that SUBSTR would make this class of question answerable from the snapshot.

Run: cd backend && python -m scripts.audit_vale_composition
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl

REPO = Path(__file__).resolve().parents[2]
WORKBOOK = REPO / "reference" / "workbook" / "Fechamento MBC 06.2026.xlsx"

MESES = {1: "Janeiro", 2: "Fevereiro", 3: "Março", 4: "Abril", 5: "Maio", 6: "Junho"}
BASE_COL = {1: 3, 2: 4, 3: 5, 4: 6, 5: 7, 6: 8}
VR_ROW, VT_ROW = 122, 123

#: Per-day vale rates, read out of the lançamento histórico in the raw extrato
#: ("Calculo: 14 dias x R$ 18,76"). Per person for VT; VR is the same for everyone.
RATES = {"VSR": 10.80, "MLA": 18.76, "JVO": 33.60}
VR_RATE = 46.10

#: The two hand-typed terms this script exists to explain. 543,22 is SOLVED (it is
#: 020.080.0050 507,10 + 020.080.0060 36,12); 35,52 is the one that survives.
ABERTOS = (35.52, 543.22)

#: Accounts that carry a separate estagiário VR/VT payable OUTSIDE the transitória.
ESTAG_VALE = ("020.080.0050", "020.080.0060")


def _brl(v: float | None) -> str:
    if v is None:
        return "—"
    s = f"{abs(v):,.2f}".replace(",", "@").replace(".", ",").replace("@", ".")
    return f"{'-' if v < 0 else ''}{s}"


def _kind(historico: str) -> str:
    h = str(historico).lower()
    if "refei" in h or "vr" in h:
        return "VR"
    return "VT"


def main() -> None:
    from dotenv import load_dotenv

    load_dotenv(REPO / "backend" / ".env")
    from app.api.providers import get_snapshot_store

    snaps = get_snapshot_store().snapshots_by_year(2026, client_id="mbc")
    base = openpyxl.load_workbook(WORKBOOK, data_only=True)["Base_Resultado Mensal_V2"]
    formulas = openpyxl.load_workbook(WORKBOOK, data_only=False)["Base_Resultado Mensal_V2"]
    months = [m for m in MESES if m in snaps]

    print("=" * 78)
    print("1. WHAT THE BOOK'S VT ROW (r123) IS MADE OF, MONTH BY MONTH")
    print("=" * 78)
    print(f"  {'mês':11}{'fórmula no livro':22}{'valor':>10}{'MLA só':>10}{'3 pessoas':>11}  leitura")
    for m in months:
        vt = {}
        for r in snaps[m].get("vale_prof") or []:
            if _kind(r.get("historico", "")) == "VT":
                vt[str(r.get("sigla"))] = round(float(r.get("valor") or 0.0), 2)
        mla = vt.get("MLA", 0.0)
        todos = round(sum(vt.values()), 2)
        book = float(base.cell(VT_ROW, BASE_COL[m]).value or 0.0)
        raw = str(formulas.cell(VT_ROW, BASE_COL[m]).value)
        if abs(book - mla) < 0.02:
            leitura = "MLA só — bate com a nossa regra"
        elif abs(book - todos) < 0.02:
            leitura = "as TRÊS pessoas — mês não ajustado"
        else:
            leitura = "NÃO é nenhum dos dois ⇒ tem termo digitado"
        print(
            f"  {MESES[m]:11}{raw:22}{_brl(book):>10}{_brl(mla):>10}{_brl(todos):>11}  {leitura}"
        )

    print("\n" + "=" * 78)
    print("2. TODO VALE É N DIAS × UMA DIÁRIA (as diárias estão no histórico do DB)")
    print("=" * 78)
    print(f"  diárias: VR {_brl(VR_RATE)}/dia · " + " · ".join(
        f"{sg} {_brl(r)}/dia" for sg, r in RATES.items()
    ))
    print(f"\n  {'mês':11}{'pessoa':7}{'tipo':5}{'valor':>10}{'dias':>8}")
    for m in months:
        for r in snaps[m].get("vale_prof") or []:
            sg = str(r.get("sigla"))
            v = round(float(r.get("valor") or 0.0), 2)
            k = _kind(r.get("historico", ""))
            rate = VR_RATE if k == "VR" else RATES.get(sg)
            dias = f"{v / rate:.2f}" if rate else "?"
            print(f"  {MESES[m]:11}{sg:7}{k:5}{_brl(v):>10}{dias:>8}")

    print("\n" + "=" * 78)
    print("3. DIAS DE VR × DIAS DE VT — o teste de consistência")
    print("=" * 78)
    print("  O VR é pago por dia trabalhado e o VT só pelos dias em que a pessoa veio;")
    print("  então a diferença de dias entre os dois deve ser pequena e estável.\n")
    print(f"  {'mês':11}{'VR dias':>9}{'VT dias (ADM)':>15}{'diferença':>11}")
    for m in months:
        vr = vt = 0.0
        for r in snaps[m].get("vale_prof") or []:
            if str(r.get("sigla")) != "MLA":
                continue
            if _kind(r.get("historico", "")) == "VR":
                vr = float(r.get("valor") or 0.0)
            else:
                vt = float(r.get("valor") or 0.0)
        dvr = vr / VR_RATE if vr else 0.0
        dvt = vt / RATES["MLA"] if vt else 0.0
        print(f"  {MESES[m]:11}{dvr:>9.0f}{dvt:>15.0f}{dvr - dvt:>+11.0f}")
    print(
        "\n  A diferença fica entre +3 e +5 em TODOS os meses. É por isso que sabemos que\n"
        "  os 14 dias de VT de janeiro estão certos: com 16 dias a diferença cairia para\n"
        "  +2, que não acontece em mês nenhum. Ou seja, o nosso 262,64 é o VT completo\n"
        "  dela em janeiro — os 35,52 são algo somado EM CIMA de um valor já correto,\n"
        "  não um pedaço que falta."
    )

    print("\n" + "=" * 78)
    print("4. OS DOIS TERMOS DIGITADOS: 35,52 E 543,22")
    print("=" * 78)
    rates_all = [VR_RATE, *RATES.values()]
    for t in ABERTOS:
        print(f"\n  {_brl(t)}:")
        hits = [
            f"{d} dias × {_brl(rate)}"
            for rate in rates_all
            for d in range(1, 26)
            if abs(d * rate - t) < 0.005
        ]
        print(f"    como N dias × diária conhecida: {hits or 'NENHUMA combinação'}")
        pares = [
            f"{d1}×{_brl(r1)} + {d2}×{_brl(r2)}"
            for r1 in rates_all
            for r2 in rates_all
            for d1 in range(26)
            for d2 in range(26)
            if (d1 or d2) and abs(d1 * r1 + d2 * r2 - t) < 0.005
        ]
        print(f"    como soma de duas diárias:      {pares[:3] or 'NENHUMA combinação'}")
        # Does the value exist as a stored number anywhere?
        found = []
        for m in sorted(snaps):
            for v in _walk(snaps[m]):
                if abs(v - t) < 0.005:
                    found.append(MESES.get(m, str(m)))
                    break
        print(f"    existe como lançamento único:   {found or 'NÃO — em nenhum dos meses'}")
        # And as the SUM of the separate estagiário VR+VT payable? This is what solves
        # 543,22, and it is the check my first pass forgot: a hand-typed term can be a
        # COMPOSITION of stored rows, not a stored row.
        for m in sorted(snaps):
            partes = [
                r
                for r in (snaps[m].get("despesas_conta") or [])
                if str(r.get("id_conta")) in ESTAG_VALE
            ]
            soma = round(sum(float(r.get("total") or 0.0) for r in partes), 2)
            if partes and abs(soma - t) < 0.005:
                print(f"    ✅ RESOLVIDO em {MESES.get(m, m)}: soma de {len(partes)} lançamentos")
                for r in partes:
                    print(
                        f"         {r['id_conta']}  {str(r.get('nome_conta'))[:18]:20}"
                        f"{_brl(float(r['total'])):>10}"
                    )

    print(
        "\n  Janeiro é o mês em que o VR foi pago por 18 dias e o VT da MLA cobriu só 14 —\n"
        "  uma diferença de 4 dias. E 35,52 = 4 × 8,88 exatamente. É uma pista, não prova:\n"
        "  8,88 não é uma diária que apareça em lugar nenhum. Continua sendo pergunta para\n"
        "  o financeiro, e é a ÚNICA coisa do bloco de vale que ainda não sabemos."
    )


def _walk(o: Any):
    if isinstance(o, dict):
        for v in o.values():
            yield from _walk(v)
    elif isinstance(o, list):
        for v in o:
            yield from _walk(v)
    elif isinstance(o, (int, float)) and not isinstance(o, bool):
        yield round(float(o), 2)


if __name__ == "__main__":
    main()
