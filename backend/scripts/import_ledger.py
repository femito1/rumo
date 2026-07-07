# backend/scripts/import_ledger.py
"""Import the hand-built per-lawyer ledger (``Base_Resultado Mensal_V2``) from a
closing workbook into the monthly SISJURI snapshots as a ``ledger`` block.

The ledger supplies per-area Custo equipe / Comissão / Despesas Equipe that
SISJURI cannot reproduce (manual per-lawyer splits) and the institutional
despesa total used to derive the per-area Despesa Institucional rateio. We read
the sheet's **cached values** by locating section anchors by label (robust to
per-month staff churn), then merge one ``ledger`` block per competence month
into that month's existing snapshot (read-modify-write), preserving all SISJURI
data. Validated to the centavo against the client dashboard (2026-07).

Usage:
    python -m scripts.import_ledger --workbook "reference/workbook/Fechamento MBC 05.2026.xlsx" \
        --client mbc --ano 2026 [--months 1,2,3,4,5] [--dry-run]
"""
from __future__ import annotations

import argparse

from app.closing.ledger_import import (
    AREA_ROW_LABELS,
    FIRST_MONTH_COL,
    LedgerMonth,
    N_MONTHS,
    despesa_institucional_rateio,
    month_reader_from_matrix,
    parse_ledger_month,
)
from app.closing.workbook_layouts import AREAS

_SHEET = "Base_Resultado Mensal_V2"


def _build_label_map(ws) -> tuple[dict[str, int], dict[str, int]]:
    """Return (label -> first row) and (area -> Despesas Área subtotal row).

    The "Despesas Área" subtotal rows share bare area labels with other blocks,
    so they are resolved positionally: the three rows directly under the
    "Despesas Área:" header, in workbook order (Contencioso, Econômico,
    Arbitragem e Compliance).
    """
    label_rows: dict[str, int] = {}
    despesas_area_header: int | None = None
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v is None:
            continue
        label = str(v).strip()
        # Keep the FIRST occurrence for the anchor labels we care about.
        if label not in label_rows:
            label_rows[label] = r
        if label == "Despesas Área:":
            despesas_area_header = r
    if despesas_area_header is None:
        raise SystemExit("Could not locate 'Despesas Área:' header in the ledger")
    da_labels: dict[str, str] = AREA_ROW_LABELS["despesas_area"]  # type: ignore[assignment]
    despesas_area_rows = {
        "Contencioso": despesas_area_header + 1,
        "Econômico": despesas_area_header + 2,
        "Arbitragem": despesas_area_header + 3,
    }
    # Sanity: the resolved rows must carry the expected bare labels.
    for area, row in despesas_area_rows.items():
        got = str(ws.cell(row=row, column=1).value or "").strip()
        want = da_labels[area]
        if got != want:
            raise SystemExit(
                f"Despesas Área row mismatch for {area}: expected {want!r} at row {row}, got {got!r}"
            )
    return label_rows, despesas_area_rows


def _has_month_data(ws, label_rows: dict[str, int], month_index: int) -> bool:
    """A competence month is present if Receita de honorários has a value."""
    receita_row = label_rows.get("Receita de honorários")
    if receita_row is None:
        return False
    v = ws.cell(row=receita_row, column=FIRST_MONTH_COL + month_index).value
    return isinstance(v, (int, float)) and v != 0


def main() -> None:
    ap = argparse.ArgumentParser(description="Import Base_Resultado ledger into snapshots.")
    ap.add_argument("--workbook", required=True, help="Path to the closing .xlsx")
    ap.add_argument("--client", default="mbc", help="client_id (default: mbc)")
    ap.add_argument("--ano", type=int, default=2026, help="competence year (default: 2026)")
    ap.add_argument(
        "--months",
        default="",
        help="comma-separated 1-based months to import (default: all present)",
    )
    ap.add_argument("--dry-run", action="store_true", help="print, do not persist")
    args = ap.parse_args()

    import openpyxl

    wb = openpyxl.load_workbook(args.workbook, data_only=True, read_only=True)
    ws = wb[_SHEET]
    label_rows, despesas_area_rows = _build_label_map(ws)

    def value_at(row: int, col: int):
        return ws.cell(row=row, column=col).value

    wanted = (
        {int(x) for x in args.months.split(",") if x.strip()}
        if args.months
        else set(range(1, N_MONTHS + 1))
    )

    parsed: dict[str, LedgerMonth] = {}
    for m in sorted(wanted):
        idx = m - 1
        if not _has_month_data(ws, label_rows, idx):
            continue
        reader = month_reader_from_matrix(
            label_rows, value_at, month_index=idx, despesas_area_rows=despesas_area_rows
        )
        parsed[f"{args.ano}-{m:02d}"] = parse_ledger_month(reader, month=m)

    if not parsed:
        raise SystemExit("No competence months with data found in the ledger.")

    print(f"Parsed ledger for {len(parsed)} month(s) from '{_SHEET}':")
    for ano_mes, lm in parsed.items():
        di = despesa_institucional_rateio(lm)
        print(f"  {ano_mes}:")
        for area in AREAS:
            print(
                f"    {area:12} custo={lm.custo_equipe.get(area, 0.0):>12,.2f}"
                f"  comissao={lm.comissao.get(area, 0.0):>10,.2f}"
                f"  desp_eq={lm.despesas_equipe.get(area, 0.0):>10,.2f}"
                f"  desp_inst={di.get(area, 0.0):>12,.2f}"
            )

    if args.dry_run:
        print("\n[dry-run] nothing persisted.")
        return

    from app.api.providers import get_snapshot_store

    store = get_snapshot_store()
    persisted = 0
    for ano_mes, lm in parsed.items():
        snap = store.get(ano_mes, client_id=args.client)
        if snap is None:
            print(f"  ! no snapshot for {ano_mes}; skipping (import SISJURI first)")
            continue
        snap["ledger"] = {
            "custo_equipe": {a: round(lm.custo_equipe.get(a, 0.0), 2) for a in AREAS},
            "comissao": {a: round(lm.comissao.get(a, 0.0), 2) for a in AREAS},
            "despesas_equipe": {a: round(lm.despesas_equipe.get(a, 0.0), 2) for a in AREAS},
            "despesa_institucional_total": round(lm.despesa_institucional_total, 2),
        }
        store.put(ano_mes, snap, client_id=args.client)
        persisted += 1
    print(f"\nPersisted ledger block into {persisted} snapshot(s).")


if __name__ == "__main__":
    main()
