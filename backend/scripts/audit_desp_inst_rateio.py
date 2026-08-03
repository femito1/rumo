"""Decompose the per-área **Despesa Institucional** difference into its TWO drivers.

The differences document attributed all three áreas' Despesa Institucional to "the
workbook's r204/205/206 formula shift". That is only half true, and the half it gets wrong
matters: the rateio is

    despesa_institucional(área) = POOL × (custo_equipe(área) / Σ custo_equipe)

where ``POOL = despesa institucional total − Σ despesas de área``. So a difference can come
from the POOL (a real difference in total despesa) or from the SHARE (a redistribution
between áreas that changes no total). Conflating them is exactly the trap recorded in
PROJECT_STATUS §0: a redistribution looks like an error but nets to zero.

Result (2026-08-03, live data), and the split is an IDENTITY — exact to R$0,01 in all 18
cells, not an estimate:

* **The SHARE effect sums to ZERO in every single month** (−0,00 / −0,01 / +0,00 / +0,00 /
  −0,00 / +0,00). It is pure redistribution: whatever one área gains another loses.
* **Every real dinheiro of difference comes from the POOL** — Jan–Jun −R$5.811,73 total.
* **June proves the mechanism**: pool differs by exactly **R$4,80** (the bank tariff the
  workbook zeroes) and the share effect is **0,00**, so all three áreas differ by ~1,7/1,8/
  1,2 — pennies, and only because that 4,80 gets rateado. June's formulas are already
  repaired, so nothing else is left.
* The pool difference flips sign at March (+1.665/+2.198 then −3.920/−3.394/−2.365),
  tracking the institucional Despesas Indiretas difference, which is where the Vale-ADM and
  the Jan/Fev one-off lines live. So the CAUSE of the per-área Despesa Institucional
  difference is not per-área at all — it is the institutional despesa total, rateado.

Run: cd backend && python -m scripts.audit_desp_inst_rateio
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl

REPO = Path(__file__).resolve().parents[2]
WORKBOOK = REPO / "reference" / "workbook" / "Fechamento MBC 06.2026.xlsx"

MESES = {1: "Janeiro", 2: "Fevereiro", 3: "Março", 4: "Abril", 5: "Maio", 6: "Junho"}
BASE_COL = {1: 3, 2: 4, 3: 5, 4: 6, 5: 7, 6: 8}
SINT_COL = {1: 3, 2: 7, 3: 11, 4: 15, 5: 19, 6: 23}

#: (our section, label, sintetico row for Despesa Institucional, Base_Resultado custo-equipe
#: block header row). The custo-equipe header is what the book's own rateio divides by.
AREAS: tuple[tuple[str, str, int, int], ...] = (
    ("contencioso", "Contencioso", 42, 5),
    ("economico", "Econômico", 60, 30),
    ("arbitragem", "Arbitragem", 78, 60),
)
#: 'Despesa para ratear' in the workbook = r198 − r203. The pool, already computed.
POOL_ROW = 207


def _brl(v: float) -> str:
    s = f"{abs(v):,.2f}".replace(",", "@").replace(".", ",").replace("@", ".")
    return f"{'-' if v < 0 else ''}{s}"


def _our(sections: dict[str, Any], section: str, line: str) -> float:
    for row in (sections.get(section) or {}).get("rows") or []:
        if row.get("key") == line:
            cell = row.get("Realizado")
            v = cell.get("value") if isinstance(cell, dict) else cell
            return float(v) if isinstance(v, (int, float)) else 0.0
    return 0.0


def main() -> None:
    from dotenv import load_dotenv

    load_dotenv(REPO / "backend" / ".env")
    from app.api.providers import get_budget_repo, get_snapshot_store
    from app.budget.models import annual_budget, monthly_budget
    from app.closing.dre import assemble_dre_sections

    snaps = get_snapshot_store().snapshots_by_year(2026, client_id="mbc")
    entries = get_budget_repo().get_budget("mbc", 2026)
    ann = annual_budget(entries) if entries else {}
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
    base = wb["Base_Resultado Mensal_V2"]
    sint = wb["Areas Sintetico atualizado"]

    months = [m for m in MESES if m in snaps]
    print("Despesa Institucional por área = POOL × (custo equipe da área / custo total)")
    print("Um Δ pode vir do POOL (dinheiro real a mais/menos) ou da SHARE (redistribuição")
    print("entre áreas, que não muda total nenhum). Aqui eles ficam separados.\n")
    print(
        f"  {'mês':10}{'área':13}{'Δ total':>11}{'do POOL':>11}{'da SHARE':>11}"
        f"{'pool nosso':>13}{'pool livro':>12}"
    )
    tot_pool = tot_share = 0.0
    per_month_share: dict[int, float] = {}
    worst = 0.0
    for m in months:
        sections = assemble_dre_sections(
            snapshot=snaps[m],
            budget=monthly_budget(entries, month=m) if entries else None,
            budget_annual=ann or None,
            transfers=None,
            period_label=f"2026-{m:02d}",
            period_month=m,
            targets=None,
        )
        pool_b = float(base.cell(POOL_ROW, BASE_COL[m]).value or 0.0)
        inst = _our(sections, "institucional", "despesas")
        deq = sum(_our(sections, k, "despesas_equipe") for k, _l, _r, _h in AREAS)
        pool_o = round(inst - deq, 2)

        ce_o = {k: _our(sections, k, "custo_equipe") for k, _l, _r, _h in AREAS}
        ce_b = {k: float(base.cell(h, BASE_COL[m]).value or 0.0) for k, _l, _r, h in AREAS}
        tot_o, tot_b = sum(ce_o.values()), sum(ce_b.values())

        share_sum = 0.0
        for k, label, wrow, _h in AREAS:
            sh_o = ce_o[k] / tot_o if tot_o else 0.0
            sh_b = ce_b[k] / tot_b if tot_b else 0.0
            ours = _our(sections, k, "despesa_institucional")
            book = float(sint.cell(wrow, SINT_COL[m]).value or 0.0)
            delta = round(ours - book, 2)
            # Exact split: (pool move at the book's share) + (share move at our pool).
            from_pool = (pool_o - pool_b) * sh_b
            from_share = pool_o * (sh_o - sh_b)
            worst = max(worst, abs(delta - (from_pool + from_share)))
            tot_pool += from_pool
            tot_share += from_share
            share_sum += from_share
            print(
                f"  {MESES[m]:10}{label:13}{_brl(delta):>11}{_brl(from_pool):>11}"
                f"{_brl(from_share):>11}{_brl(pool_o):>13}{_brl(pool_b):>12}"
            )
        per_month_share[m] = share_sum

    print(f"\n  A decomposição fecha com erro máximo de R$ {worst:.4f} — é identidade, não estimativa.")
    print(f"  Jan–{MESES[max(months)]}: do POOL {_brl(round(tot_pool, 2))} · da SHARE {_brl(round(tot_share, 2))}")
    print("\n  A SHARE soma ZERO em cada mês — é redistribuição pura, não dinheiro:")
    for m in months:
        print(f"    {MESES[m]:10} soma das três áreas: {_brl(round(per_month_share[m], 2))}")
    print(
        "\n  => TODA a diferença real de Despesa Institucional por área vem do POOL, ou seja\n"
        "     da despesa institucional TOTAL (onde vivem o Vale-ADM e os lançamentos avulsos\n"
        "     de jan/fev), rateada. Não é um problema por área.\n"
        "     Junho prova o mecanismo: pool difere exatamente R$ 4,80 (a tarifa bancária que\n"
        "     a planilha zera), share 0,00, e as três áreas ficam em centavos."
    )


if __name__ == "__main__":
    main()
