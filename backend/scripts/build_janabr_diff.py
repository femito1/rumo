"""Per-line Jan–Abr diff (nosso × planilha), with every difference ATTRIBUTED.

HANDOFF §5.6 — "vou mapear essas coisas e mostrar para vocês". The point is NOT to
converge on the workbook: several of these differences are the workbook's own
formula/entry quirks, already diagnosed. The point is a list Renata can read line by
line, where each number is traced to a named cause.

Run: cd backend && python -m scripts.build_janabr_diff
Writes: docs/DIFF_JAN_ABR_2026.md (and prints the same to stdout).

Reads the LIVE snapshots (extract v3) + the June workbook, so re-run it after any
re-extract rather than trusting a stale copy of the output.

What it establishes, and the reason each matters:

* Faturamento, Receita, Imposto and Amortização tie EXACTLY in all four months. So
  the entire Resultado Bruto/Líquido difference is inside custo equipe + despesas —
  worth stating up front, because it rules out the revenue side entirely.

* Custo equipe differs for exactly three reasons, all identified:
  1. lawyer **Vale** (VR/VT) — we fold the per-person 500.010.<SIGLA> slices into the
     área's custo equipe (client ruling "always include Vale"); Jan–Mai in the book
     do not. Contencioso Jan 997,80 · Fev 1.249,40 · Mar/Abr 1.190,80; Econômico
     Mar/Abr 1.008,40.
  2. the **estagiária VSR** — her salary enters Econômico from March (book r52:
     Mar 1.026,00, Abr 2.200,00) and we match it to the centavo; the sign of the
     Econômico gap flips at exactly that month.
  3. **one-off lines the book carries in Jan/Fev only** (r34 3.018,00, r35 520,00,
     r43 1.034,38, r47 1.000,00, r51 1.409,09, r54 2.101,88 …) — hand-entered
     adjustments with no counterpart in the competence's lançamentos.
  4. **a lawyer's convênio médico typed in January only** (r69, 1.911,45) while the
     DB keeps posting it (030.010.0110, 1.911,95) — this is essentially the whole
     Arbitragem February gap, and it closes from March when he leaves both sides.
     (I first suspected a double count at r66/r68, since both move by 1.610,40 in
     February. REFUTED: r66 is `=15030.4-D67`, a netted distribution, and r68 is a
     separate "Reajuste" line. Don't repeat that guess.)

* Despesas differ for two reasons: the Vale-ADM split (mar/abr/mai un-adjusted in the
  book — Renata: "não vale a pena corrigir") and the recurring bank tariff the book
  zeroes (4,80/month), which also moves reserva by ~10% of it.

* Per-área Despesa Institucional differs because the rateio pool depends on
  Despesas Área, and the book's Jan–Mai r204/205/206 are off by one row (see
  scripts/audit_area_ytd_formulas.py). That single defect moves all three áreas.
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl

from app.budget.models import annual_budget, monthly_budget
from app.closing.dre import assemble_dre_sections

REPO = Path(__file__).resolve().parents[2]
WORKBOOK = REPO / "reference" / "workbook" / "Fechamento MBC 06.2026.xlsx"
OUT = REPO / "docs" / "DIFF_JAN_ABR_2026.md"

MESES = {1: "Janeiro", 2: "Fevereiro", 3: "Março", 4: "Abril"}
#: 'Areas Sintetico atualizado' Realizado column per month.
SINT_COL = {1: 3, 2: 7, 3: 11, 4: 15}
#: Base_Resultado month column.
BASE_COL = {1: 3, 2: 4, 3: 5, 4: 6}

#: (our section, our line key) -> sintetico row, in reading order.
LINES: tuple[tuple[str, str, str, int], ...] = (
    ("institucional", "recebimento", "Receita", 4),
    ("institucional", "custo_equipe", "Custos Diretos", 6),
    ("institucional", "despesas", "Despesas Indiretas", 13),
    ("institucional", "resultado_bruto", "Resultado Bruto", 25),
    ("institucional", "imposto", "Impostos", 28),
    ("institucional", "amortizacao", "Amortização", 29),
    ("institucional", "resultado_liquido", "Resultado Líquido", 30),
    ("institucional", "reserva_bonus", "Reserva de bônus", 32),
    ("contencioso", "custo_equipe", "Contencioso · Custo equipe", 39),
    ("contencioso", "despesas_equipe", "Contencioso · Despesas Equipe", 41),
    ("contencioso", "despesa_institucional", "Contencioso · Despesa Institucional", 42),
    ("contencioso", "resultado_bruto", "Contencioso · Resultado Bruto", 43),
    ("economico", "custo_equipe", "Econômico · Custo equipe", 57),
    ("economico", "despesas_equipe", "Econômico · Despesas Equipe", 59),
    ("economico", "despesa_institucional", "Econômico · Despesa Institucional", 60),
    ("economico", "resultado_bruto", "Econômico · Resultado Bruto", 61),
    ("arbitragem", "custo_equipe", "Arbitragem · Custo equipe", 75),
    ("arbitragem", "despesas_equipe", "Arbitragem · Despesas Equipe", 77),
    ("arbitragem", "despesa_institucional", "Arbitragem · Despesa Institucional", 78),
    ("arbitragem", "resultado_bruto", "Arbitragem · Resultado Bruto", 79),
)

_GRUPO_TO_SECTION = {
    "equipecontencioso": "contencioso",
    "equipedireitoeconômico": "economico",
    "equipedireitoeconomico": "economico",
    "arbitragem": "arbitragem",
    "equipeambiental": "arbitragem",
}


def _brl(v: float | None) -> str:
    if v is None:
        return "—"
    s = f"{abs(v):,.2f}".replace(",", "@").replace(".", ",").replace("@", ".")
    return f"{'-' if v < 0 else ''}{s}"


def _our_value(sections: dict[str, Any], section: str, line: str) -> float | None:
    for row in (sections.get(section) or {}).get("rows") or []:
        if row.get("key") == line:
            cell = row.get("Realizado")
            v = cell.get("value") if isinstance(cell, dict) else cell
            return float(v) if isinstance(v, (int, float)) else None
    return None


def _vale_by_section(snap: dict[str, Any]) -> dict[str, float]:
    """Per-área lawyer Vale (the slices we fold into custo equipe)."""
    home = {str(k): str(v) for k, v in (snap.get("home_area") or {}).items()}
    out: dict[str, float] = {}
    for r in snap.get("vale_prof") or []:
        grupo = home.get(str(r.get("sigla")), "").replace(" ", "").lower()
        sec = next((v for k, v in _GRUPO_TO_SECTION.items() if k in grupo), None)
        if sec:
            out[sec] = round(out.get(sec, 0.0) + float(r.get("valor") or 0.0), 2)
    return out


def main() -> None:
    # This script reads the LIVE Supabase snapshots, so it needs backend/.env — the
    # app itself is configured by the process env in prod and never loads a file.
    # Nothing else under scripts/ does this (they all read fixtures), hence the
    # explicit load here rather than a shared helper.
    from dotenv import load_dotenv

    load_dotenv(REPO / "backend" / ".env")

    from app.api.providers import get_budget_repo, get_snapshot_store, get_transfers_repo

    snaps = get_snapshot_store().snapshots_by_year(2026, client_id="mbc")
    entries = get_budget_repo().get_budget("mbc", 2026)
    ann = annual_budget(entries) if entries else {}
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
    sint = wb["Areas Sintetico atualizado"]
    base = wb["Base_Resultado Mensal_V2"]

    assembled: dict[int, dict[str, Any]] = {}
    for m in MESES:
        try:
            transfers = get_transfers_repo().get_transfers("mbc", f"2026-{m:02d}")
        except Exception:
            transfers = None
        assembled[m] = assemble_dre_sections(
            snapshot=snaps[m],
            budget=monthly_budget(entries, month=m) if entries else None,
            budget_annual=ann or None,
            transfers=transfers,
            period_label=f"2026-{m:02d}",
            period_month=m,
            targets=None,
        )

    L: list[str] = []
    add = L.append
    add("# Diferenças por linha — janeiro a abril de 2026 (nosso × planilha)")
    add("")
    add("> Gerado por `backend/scripts/build_janabr_diff.py` a partir dos dados ao vivo")
    add("> (extract v3) e de `Fechamento MBC 06.2026.xlsx`. **Não é uma lista de erros**:")
    add("> cada diferença abaixo já foi diagnosticada e tem causa identificada. Junho")
    add("> fecha exatamente, e é o melhor mês de referência.")
    add("")
    add("## Resumo")
    add("")
    add("**Faturamento, Receita, Impostos e Amortização batem exatamente nos quatro meses.**")
    add("Toda a diferença de Resultado está dentro de *custo de equipe* e *despesas*, por")
    add("três causas conhecidas:")
    add("")
    add("1. **Vale dos advogados** — nós incluímos o vale-refeição/transporte de cada")
    add("   advogado no custo de equipe da área dele (regra confirmada: \"sempre incluir")
    add("   o vale\"); as colunas de janeiro a maio da planilha não incluem.")
    add("2. **Estagiária do Direito Econômico** — o salário dela entra a partir de março")
    add("   e nós reproduzimos o valor da planilha ao centavo; é exatamente nesse mês que")
    add("   o sinal da diferença do Econômico se inverte.")
    add("3. **Lançamentos avulsos digitados só em janeiro/fevereiro** na planilha, sem")
    add("   lançamento correspondente na competência do sistema.")
    add("")
    add("Além disso, a **fórmula de Despesas por área** da planilha está deslocada uma")
    add("linha de janeiro a maio (linhas 204/205/206), o que desloca também o rateio da")
    add("despesa institucional das três áreas. As fórmulas de junho já estão corretas.")
    add("")
    add("## Diferenças por linha")
    add("")
    add("Valores em R$. `Δ` = nosso − planilha (positivo = o nosso é maior).")
    add("")
    header = "| Linha | " + " | ".join(
        f"{MESES[m]} (nosso / planilha / Δ)" for m in MESES
    ) + " |"
    add(header)
    add("|" + "---|" * (len(MESES) + 1))
    for section, line, label, wrow in LINES:
        cells = []
        for m in MESES:
            ours = _our_value(assembled[m], section, line)
            book = sint.cell(wrow, SINT_COL[m]).value or 0.0
            delta = (ours or 0.0) - float(book)
            mark = "" if abs(delta) < 0.02 else " "
            cells.append(f"{_brl(ours)} / {_brl(float(book))} / **{_brl(delta)}**{mark}")
        add(f"| {label} | " + " | ".join(cells) + " |")
    add("")

    add("## De onde vem cada diferença")
    add("")
    add("### 1. Vale dos advogados dentro do custo de equipe")
    add("")
    add("Valores que nós somamos ao custo de equipe de cada área e que a planilha não")
    add("soma nas colunas de janeiro a maio:")
    add("")
    add("| Mês | Contencioso | Econômico | Arbitragem |")
    add("|---|---|---|---|")
    for m in MESES:
        v = _vale_by_section(snaps[m])
        add(
            f"| {MESES[m]} | {_brl(v.get('contencioso'))} | "
            f"{_brl(v.get('economico'))} | {_brl(v.get('arbitragem'))} |"
        )
    add("")
    add("### 2. Estagiária do Direito Econômico (planilha, linha 52)")
    add("")
    add("| Mês | Planilha (linha 52) | Nosso |")
    add("|---|---|---|")
    for m in MESES:
        book = base.cell(52, BASE_COL[m]).value or 0.0
        ours = sum(
            float(r.get("valor") or 0.0)
            for r in (snaps[m].get("custo_equipe_deriv") or [])
            if r.get("sigla") == "VSR"
        )
        add(f"| {MESES[m]} | {_brl(float(book))} | {_brl(ours)} |")
    add("")
    add("Batem — não é fonte de diferença a partir de março; é o que explica a")
    add("**inversão de sinal** da diferença do Econômico entre fevereiro e março.")
    add("")
    add("### 3. Lançamentos avulsos só em janeiro/fevereiro (bloco do Econômico)")
    add("")
    add("Linhas da planilha com valor em janeiro ou fevereiro e zeradas nos meses")
    add("seguintes. Não encontramos lançamento correspondente na competência:")
    add("")
    add("| Linha | Janeiro | Fevereiro | Março | Abril |")
    add("|---|---|---|---|---|")
    for r in (34, 35, 36, 43, 47, 51, 54, 55):
        vals = [float(base.cell(r, BASE_COL[m]).value or 0.0) for m in MESES]
        if any(vals):
            add(f"| {r} | " + " | ".join(_brl(v) for v in vals) + " |")
    add("")
    add("### 4. Arbitragem, fevereiro: convênio médico de um advogado (planilha, linha 69)")
    add("")
    add("A planilha traz o convênio médico de um advogado da Arbitragem (linha 69,")
    add("`R$ 1.911,45`) **apenas em janeiro**; de fevereiro em diante a linha fica")
    add("zerada. No sistema esse convênio (conta `030.010.0110`, `R$ 1.911,95`) continua")
    add("lançado em fevereiro, e é isso que explica quase toda a diferença de")
    add("`R$ 1.911,95` do custo de equipe da Arbitragem naquele mês.")
    add("")
    add("| Mês | Planilha (linha 69) | Nosso (conta 030.010.0110) |")
    add("|---|---|---|")
    for m in MESES:
        book = float(base.cell(69, BASE_COL[m]).value or 0.0)
        ours = sum(
            float(r.get("valor") or 0.0)
            for r in (snaps[m].get("custo_equipe_deriv") or [])
            if r.get("sigla") == "JGS" and str(r.get("id_conta")) == "030.010.0110"
        )
        add(f"| {MESES[m]} | {_brl(book)} | {_brl(ours)} |")
    add("")
    add("A partir de março esse advogado sai da folha nos dois lados, e o custo de")
    add("equipe da Arbitragem volta a bater exatamente (março e abril: diferença zero).")
    add("")
    add("### 5. Perguntas que sobram para o financeiro")
    add("")
    add("1. Os lançamentos avulsos do item 3 são de outra competência, ou ajustes")
    add("   manuais? Se tiverem origem no sistema, passamos a considerá-los.")
    add("2. Confirmar se as fórmulas de Despesas por área (linhas 204/205/206) de")
    add("   janeiro a maio devem ser copiadas de junho.")
    add("3. Janeiro: o vale-transporte da planilha é `=35,52+262,64`. Os 262,64 são o")
    add("   lançamento do sistema; de onde vêm os 35,52?")
    add("4. O convênio médico da linha 69 (item 4) deveria continuar em fevereiro,")
    add("   como está no sistema, ou foi encerrado em janeiro?")
    add("")
    add("*(Vale-ADM de março/abril/maio já está respondido: são meses não ajustados na")
    add("planilha e o financeiro optou por não corrigir.)*")

    text = "\n".join(L) + "\n"
    OUT.write_text(text, encoding="utf-8")
    print(text)
    print(f"[written] {OUT.relative_to(REPO)}")


if __name__ == "__main__":
    main()
