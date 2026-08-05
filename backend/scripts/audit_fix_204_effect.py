"""What would repairing the workbook's r204/205/206 actually change?

Asked directly (2026-08-04): "what would be the effect of just fixing the 204/205/206 in
the workbook?" Measured rather than reasoned about, because the repair is NOT local — those
three rows feed four things:

  * per-área Despesas Equipe directly (sintetico r41/r59/r77 read base r204/205/206);
  * ``r203 = 204+205+206``, and the rateio pool ``H2 = r198 − 204 − 205 − 206``;
  * per-área Despesa Institucional = pool × (área custo share)  (Rateio Mensal I2/I3/I4);
  * therefore per-área Resultado Bruto.

ANSWER, and it is narrower than it looks:

* **The headline gap does NOT move.** ``r198`` (institucional Despesas Indiretas) does not
  reference 204–206, so the institucional total — and the −R$5.003,04 Resultado Bruto YTD
  difference — is completely unchanged. The repair fixes WHICH ÁREA carries which despesa,
  nothing else. Do not let anyone expect the totals to converge afterwards.
* Per-área **Despesas Equipe** absolute error falls 10.216,30 → 4.244,76 (**−58%**) and
  **Despesa Institucional** 13.549,64 → 8.956,70 (**−34%**).
* Per-área **Resultado Bruto** barely moves (−1%): it is the sum of those two lines and the
  corrections partly offset within each área.

⚠ TRAP I FELL INTO FIRST: applying the "fix" to all six months made Resultado Bruto look 4%
WORSE. June's r204/205/206 are ALREADY the repaired formula, so overwriting them corrupts a
month that was fine. The simulation must skip June (``rows_fix = ATUAL if m == 6``). A model
that does not match the sheet produces confident nonsense.

Run: cd backend && python -m scripts.audit_fix_204_effect
"""
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parents[2]
WB = REPO / "reference" / "workbook" / "Fechamento MBC 06.2026.xlsx"
MESES = {1: "Jan", 2: "Fev", 3: "Mar", 4: "Abr", 5: "Mai", 6: "Jun"}
BASE = {1: 3, 2: 4, 3: 5, 4: 6, 5: 7, 6: 8}
SINT = {1: 3, 2: 7, 3: 11, 4: 15, 5: 19, 6: 23}

# The rows each area's Despesas-Equipe formula adds: as typed (Jan-Mai) vs repaired (Jun).
ATUAL = {
    "contencioso": [125, 129, 140, 144, 148, 152, 156, 160],
    "economico": [126, 130, 141, 145, 149, 153, 157, 161],
    "arbitragem": [127, 131, 139, 143, 147, 151, 155, 159],
}
CERTO = {
    "contencioso": [125, 129, 139, 143, 147, 151, 155, 160],
    "economico": [126, 130, 140, 144, 148, 152, 156, 161],
    "arbitragem": [127, 131, 138, 142, 146, 150, 154, 159],
}
# custo-equipe block header per area (the rateio share basis) and sintetico rows.
CUSTO_ROW = {"contencioso": 5, "economico": 30, "arbitragem": 60}
SINT_DEQ = {"contencioso": 41, "economico": 59, "arbitragem": 77}
SINT_DI = {"contencioso": 42, "economico": 60, "arbitragem": 78}
SINT_RB = {"contencioso": 43, "economico": 61, "arbitragem": 79}


def brl(v):
    s = f"{abs(v):,.2f}".replace(",", "@").replace(".", ",").replace("@", ".")
    return f"{'-' if v < 0 else ''}{s}"


def main():
    from dotenv import load_dotenv

    load_dotenv(REPO / "backend" / ".env")

    from app.api.providers import get_budget_repo, get_snapshot_store
    from app.budget.models import annual_budget, monthly_budget
    from app.closing.dre import assemble_dre_sections, convenio_mbc_shares

    snaps = get_snapshot_store().snapshots_by_year(2026, client_id="mbc")
    entries = get_budget_repo().get_budget("mbc", 2026)
    ann = annual_budget(entries) if entries else {}
    shares = convenio_mbc_shares(snaps)
    _wbv = openpyxl.load_workbook(WB, data_only=True)
    base = _wbv["Base_Resultado Mensal_V2"]
    sint = _wbv["Areas Sintetico atualizado"]

    def ours(m, sec, key):
        a = assemble_dre_sections(
            snapshot=snaps[m],
            budget=monthly_budget(entries, month=m) if entries else None,
            budget_annual=ann or None, transfers=None,
            period_label=f"2026-{m:02d}", period_month=m, targets=None,
            convenio_shares=shares,
        )
        for r in (a.get(sec) or {}).get("rows") or []:
            if r.get("key") == key:
                c = r.get("Realizado")
                v = c.get("value") if isinstance(c, dict) else c
                return float(v) if isinstance(v, (int, float)) else 0.0
        return 0.0

    tot = {k: [0.0, 0.0, 0.0] for k in ("deq", "di", "rb")}
    print(f"{'mes':5}{'area':13}{'linha':22}{'Δ hoje':>12}{'Δ se corrigir':>15}")
    for m in MESES:
        col = BASE[m]
        r198 = float(base.cell(198, col).value or 0.0)
        custo = {a: float(base.cell(CUSTO_ROW[a], col).value or 0.0) for a in ATUAL}
        tcusto = sum(custo.values())
        deq_now = {a: sum(float(base.cell(r, col).value or 0.0) for r in ATUAL[a]) for a in ATUAL}
        rows_fix = ATUAL if m == 6 else CERTO   # June already carries the repaired formula
        deq_fix = {a: sum(float(base.cell(r, col).value or 0.0) for r in rows_fix[a]) for a in rows_fix}
        pool_now = r198 - sum(deq_now.values())
        pool_fix = r198 - sum(deq_fix.values())
        for a in ATUAL:
            sh = custo[a] / tcusto if tcusto else 0.0
            di_now, di_fix = pool_now * sh, pool_fix * sh
            o_deq, o_di = ours(m, a, "despesas_equipe"), ours(m, a, "despesa_institucional")
            o_rb = ours(m, a, "resultado_bruto")
            # Workbook RB moves by exactly -(Δdeq + Δdi) when only these change.
            rb_shift = -((deq_fix[a] - deq_now[a]) + (di_fix - di_now))
            d_deq_now, d_deq_fix = o_deq - deq_now[a], o_deq - deq_fix[a]
            d_di_now, d_di_fix = o_di - di_now, o_di - di_fix
            b_rb = float(sint.cell(SINT_RB[a], SINT[m]).value or 0.0)
            d_rb_now = o_rb - b_rb
            d_rb_fix = o_rb - (b_rb + rb_shift)
            for k, now, fix in (
                ("deq", d_deq_now, d_deq_fix),
                ("di", d_di_now, d_di_fix),
                ("rb", d_rb_now, d_rb_fix),
            ):
                tot[k][0] += abs(now)
                tot[k][1] += abs(fix)
                tot[k][2] += fix
            print(f"{MESES[m]:5}{a:13}{'Despesas Equipe':22}{brl(d_deq_now):>12}{brl(d_deq_fix):>15}")
            print(f"{'':5}{'':13}{'Despesa Instituc.':22}{brl(d_di_now):>12}{brl(d_di_fix):>15}")
            print(f"{'':5}{'':13}{'Resultado Bruto':22}{brl(d_rb_now):>12}{brl(d_rb_fix):>15}")

    print()
    print("SOMA DOS |Δ| (18 células cada):")
    for k, nome in (("deq", "Despesas Equipe"), ("di", "Despesa Institucional"), ("rb", "Resultado Bruto")):
        a, b = tot[k][0], tot[k][1]
        print(f"  {nome:24} hoje {brl(a):>12}   corrigido {brl(b):>12}   ({100*(1-b/a):+.0f}%)")
    print()
    print("Δ ACUMULADO (com sinal) depois da correção:")
    for k, nome in (("deq", "Despesas Equipe"), ("di", "Despesa Institucional"), ("rb", "Resultado Bruto")):
        print(f"  {nome:24} {brl(tot[k][2]):>12}")


if __name__ == "__main__":
    main()
