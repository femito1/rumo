"""Decompose institucional **Despesas Indiretas** to a named cause per family per month.

The differences document carried this line as a single YTD number with a prose cause. That
was the last big line nobody had decomposed mechanically, and doing it (2026-08-04) closed
it almost completely: **the residual is R$0,00 in four of six months**, and every surviving
centavo has a named account behind it.

The method is the one that works on this data: never compare a total, always decompose to
the workbook's own family header rows (`Base_Resultado Mensal_V2`) and then to leaves. A
family total can tie while its leaves are wrong, and a family can look wrong while the money
is merely filed under a neighbouring label.

What it found, in order of size
-------------------------------
1. **Vale-ADM, −R$ 7.257,62 YTD — the workbook's own basis is INCONSISTENT.** Ours is
   MLA-only in every month (Renata's 2026-07-30 ruling: JVO/VSR are estagiários of the
   áreas). The book's rows 122/123 are MLA-only in fev/jun, **all three people** in abr,
   and NEITHER in jan/mar/mai. So this is not us versus a rule — it is the book applying
   three different rules across six months. Nothing to converge on; ours is the consistent
   one. (mar/abr/mai are the months Renata said *"não vale a pena corrigir"*.)
2. **Seguros, +R$ 2.539,84 in January — a lump annual premium.** `020.060.0040` posts
   **2.722,55** in January (and again in July) while the book types a flat **182,71** every
   month. `2.722,55 − 182,71 = 2.539,84`, the January Ocupação delta to the centavo. The
   book is smoothing an annual premium; the DB posts it when it is paid.
3. **The estagiária's separate vale, +R$ 543,22 in March.** `020.080.0050` VR 507,10 +
   `020.080.0060` VT 36,12 — a real payable OUTSIDE the transitória (already documented in
   `docs/SISJURI_DB.md`). It is the exact offset that made the Salários-Administração family
   look like it under-explained the vale.
4. **Two accounts the workbook has no row for at all** — the same shape as the Associações
   case, and the same conclusion (*"é sempre o banco tendo mais informação que a planilha"*):
   January `020.050.0070` **IR Fonte - ADM 169,52** and February `020.050.0160` **Relatórios
   trabalhistas - e-Social 1.032,35**. Single postings, real, simply absent from the sheet.
5. **Endomarketing ↔ Investimentos em Prospecção is presentation-only.** Individually they
   are +2.484,46 / −2.522,37 YTD; **summed they are 0,00 in four months** and the residue is
   items (3)/(4) leaking through. Both rows feed r198, so the swap cannot move a total.

After (1)–(4), the Salários Administração family residual is **0,00 in mar/abr/mai/jun** and
the only survivors are jan 169,52 and fev 1.032,35 — which are (4).

Run: cd backend && python -m scripts.audit_despesas_indiretas
"""
from __future__ import annotations

from pathlib import Path

REPO = Path(__file__).resolve().parents[2]
WORKBOOK = REPO / "reference" / "workbook" / "Fechamento MBC 06.2026.xlsx"
BASE_COL = {1: 3, 2: 4, 3: 5, 4: 6, 5: 7, 6: 8}
MESES = {1: "Janeiro", 2: "Fevereiro", 3: "Março", 4: "Abril", 5: "Maio", 6: "Junho"}

#: Our family label -> the workbook's family HEADER row in 'Base_Resultado Mensal_V2'.
#: Read off the sheet, not guessed; these are the rows whose sum IS r198.
FAMILIES = {
    "Ocupação": 85,
    "Telecomunicações": 92,
    "Despesas Gerais": 95,
    "Consultoria": 110,
    "Salários Administração": 116,
    "Administrativas": 124,
    "Investimentos em Prospecção": 137,
    "Gestão do Conhecimento": 158,
    "Endomarketing": 164,
    "Informática": 180,
}
#: The book's two ADM-vale rows, and the estagiária's own accounts.
VALE_ROWS = (122, 123)
ESTAGIARIA_LEAVES = ("Vale Refeição", "Vale Transporte")


def _brl(v: float) -> str:
    s = f"{abs(v):,.2f}".replace(",", "@").replace(".", ",").replace("@", ".")
    return f"{'-' if v < 0 else ''}{s}"


def main() -> None:
    import openpyxl
    from dotenv import load_dotenv

    load_dotenv(REPO / "backend" / ".env")
    from app.api.providers import get_budget_repo, get_snapshot_store
    from app.budget.models import annual_budget, monthly_budget
    from app.closing.dre import assemble_dre_sections, convenio_mbc_shares, is_adm_grupo

    snaps = get_snapshot_store().snapshots_by_year(2026, client_id="mbc")
    entries = get_budget_repo().get_budget("mbc", 2026)
    ann = annual_budget(entries) if entries else {}
    shares = convenio_mbc_shares(snaps)
    base = openpyxl.load_workbook(WORKBOOK, data_only=True)["Base_Resultado Mensal_V2"]

    months = [m for m in MESES if m in snaps]
    asm = {
        m: assemble_dre_sections(
            snapshot=snaps[m],
            budget=monthly_budget(entries, month=m) if entries else None,
            budget_annual=ann or None,
            transfers=None,
            period_label=f"2026-{m:02d}",
            period_month=m,
            targets=None,
            convenio_shares=shares,
        )
        for m in months
    }

    def ours(m: int, family: str) -> float:
        return next(
            (r.get("Realizado", {}).get("value") or 0.0
             for r in asm[m]["institucional"]["rows"] if r.get("Linha") == family),
            0.0,
        )

    def leaf(m: int, family_prefix: str, name: str) -> float:
        for r in asm[m]["institucional"]["rows"]:
            key = str(r.get("key") or "")
            if key.startswith(f"acct::{family_prefix}") and key.split("::")[-1] == name:
                return r.get("Realizado", {}).get("value") or 0.0
        return 0.0

    print("1. DIFERENÇA POR FAMÍLIA (nossa − planilha)")
    header = "".join(f"{MESES[m][:3]:>10}" for m in months)
    print(f"   {'família':<30}{header}{'YTD':>11}")
    grand = 0.0
    for family, row in FAMILIES.items():
        ds = [round(ours(m, family) - float(base.cell(row, BASE_COL[m]).value or 0.0), 2)
              for m in months]
        grand += sum(ds)
        print(f"   {family:<30}" + "".join(f"{_brl(d):>10}" for d in ds) + f"{_brl(round(sum(ds), 2)):>11}")
    print(f"   {'TOTAL':<30}" + " " * (10 * len(months)) + f"{_brl(round(grand, 2)):>11}")

    print("\n2. VALE-ADM — a planilha usa uma BASE DIFERENTE em cada mês")
    print(f"   {'mês':<11}{'nosso (MLA)':>13}{'planilha':>12}{'MLA?':>7}{'os três?':>10}")
    vale_total = 0.0
    for m in months:
        home = snaps[m].get("home_area") or {}
        rows = snaps[m].get("vale_prof") or []
        mla = round(sum(float(r["valor"]) for r in rows
                        if is_adm_grupo(home.get(str(r.get("sigla"))))), 2)
        todos = round(sum(float(r["valor"]) for r in rows), 2)
        book = round(sum(float(base.cell(r, BASE_COL[m]).value or 0.0) for r in VALE_ROWS), 2)
        vale_total += mla - book
        print(f"   {MESES[m]:<11}{_brl(mla):>13}{_brl(book):>12}"
              f"{('sim' if abs(book - mla) < 0.02 else '-'):>7}"
              f"{('sim' if abs(book - todos) < 0.02 else '-'):>10}")
    print(f"   YTD {_brl(round(vale_total, 2))} — e note que a planilha só bate com a NOSSA")
    print("   regra em fev/jun, usa os três em abril, e nenhuma das duas em jan/mar/mai.")

    print("\n3. SALÁRIOS ADMINISTRAÇÃO — resíduo depois de tirar vale e estagiária")
    print(f"   {'mês':<11}{'Δ família':>12}{'Δ vale':>12}{'estagiária':>12}{'resíduo':>11}")
    for m in months:
        home = snaps[m].get("home_area") or {}
        rows = snaps[m].get("vale_prof") or []
        mla = round(sum(float(r["valor"]) for r in rows
                        if is_adm_grupo(home.get(str(r.get("sigla"))))), 2)
        book = round(sum(float(base.cell(r, BASE_COL[m]).value or 0.0) for r in VALE_ROWS), 2)
        fam = round(ours(m, "Salários Administração")
                    - float(base.cell(116, BASE_COL[m]).value or 0.0), 2)
        est = round(sum(leaf(m, "Salários", n) for n in ESTAGIARIA_LEAVES), 2)
        print(f"   {MESES[m]:<11}{_brl(fam):>12}{_brl(round(mla - book, 2)):>12}"
              f"{_brl(est):>12}{_brl(round(fam - (mla - book) - est, 2)):>11}")
    print("   Os resíduos de jan/fev são duas contas que a planilha não tem: IR Fonte - ADM")
    print("   (020.050.0070, 169,52) e Relatórios trabalhistas - e-Social (020.050.0160,")
    print("   1.032,35). Lançamentos reais, únicos, ausentes do Excel.")

    print("\n4. ENDOMARKETING ↔ INVESTIMENTOS EM PROSPECÇÃO — só apresentação")
    for m in months:
        o = ours(m, "Endomarketing") + ours(m, "Investimentos em Prospecção")
        b = float(base.cell(164, BASE_COL[m]).value or 0.0) + float(
            base.cell(137, BASE_COL[m]).value or 0.0)
        print(f"   {MESES[m]:<11}{_brl(round(o, 2)):>13}{_brl(round(b, 2)):>12}"
              f"{_brl(round(o - b, 2)):>11}")
    print("   As duas linhas entram na r198, então a troca não pode mover total nenhum.")

    print("\n5. SEGUROS (020.060.0040) — prêmio anual, não mensal")
    for m in sorted(snaps):
        for r in snaps[m].get("despesas_conta") or []:
            if r.get("id_conta") == "020.060.0040":
                print(f"   2026-{m:02d} {_brl(float(r['total'])):>12}  (n={r.get('n')})")
    print("   A planilha digita 182,71 todo mês; o banco lança 2.722,55 em janeiro (e de")
    print("   novo em julho). 2.722,55 − 182,71 = 2.539,84 = a diferença de Ocupação de")
    print("   janeiro, ao centavo.")


if __name__ == "__main__":
    main()
